from __future__ import annotations

from pathlib import Path

from openpyxl.cell.cell import MergedCell
from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries

from .config import SheetTemplateConfig


def _iter_range_cells(worksheet, range_string: str):
    min_col, min_row, max_col, max_row = range_boundaries(range_string)
    yield from worksheet.iter_rows(
        min_row=min_row,
        max_row=max_row,
        min_col=min_col,
        max_col=max_col,
    )


def _row_bounds(range_string: str) -> tuple[int, int]:
    _, min_row, _, max_row = range_boundaries(range_string)
    return min_row, max_row


def _unhide_rows(worksheet, start_row: int, end_row: int) -> None:
    for row_index in range(start_row, end_row + 1):
        worksheet.row_dimensions[row_index].hidden = False


def _clear_values(worksheet, range_string: str) -> None:
    for row in _iter_range_cells(worksheet, range_string):
        for cell in row:
            if isinstance(cell, MergedCell):
                continue
            cell.value = None


def _clear_formula_values(worksheet, range_string: str) -> None:
    for row in _iter_range_cells(worksheet, range_string):
        for cell in row:
            if isinstance(cell, MergedCell):
                continue
            if isinstance(cell.value, str) and cell.value.startswith("="):
                cell.value = None


def build_sanitized_workbook(
    source_path: Path,
    output_path: Path,
    configs: tuple[SheetTemplateConfig, ...],
) -> None:
    workbook = load_workbook(source_path, keep_vba=True)

    for config in configs:
        worksheet = workbook[config.title]
        worksheet.sheet_view.showGridLines = False
        worksheet.print_area = config.visible_range
        _clear_formula_values(worksheet, config.visible_range)

        for clear_range in config.clear_ranges:
            _clear_values(worksheet, clear_range)
            start_row, end_row = _row_bounds(clear_range)
            _unhide_rows(worksheet, start_row, end_row)

        for cell_ref in config.clear_cells:
            worksheet[cell_ref].value = None

        for start_row, end_row in config.unhide_row_spans:
            _unhide_rows(worksheet, start_row, end_row)

    workbook.save(output_path)
