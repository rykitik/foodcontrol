from __future__ import annotations

from copy import copy

from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Alignment, Border
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.utils.cell import range_boundaries
from openpyxl.worksheet.worksheet import Worksheet

from app.utils.report_generator import russian_month_name

from .metadata_values import parse_decimal_metadata_value
from .template_config import (
    ColumnWidthContract,
    CostCalculationTemplateConfig,
    CostStatementTemplateConfig,
    MealSheetTemplateConfig,
)
from .template_signatures import apply_prepared_by_binding


def populate_cost_statement_worksheet(
    worksheet: Worksheet,
    payload: dict,
    config: CostStatementTemplateConfig,
) -> None:
    _assert_capacity(len(payload["rows"]), config.capacity, "ведомости")
    _apply_column_width_contract(worksheet, config.column_width_contract)
    _merge_cost_statement_requisite_text_fields(worksheet)

    worksheet[config.month_cell] = russian_month_name(payload["period_start"])
    worksheet[config.year_cell] = _year_label(payload["year"])
    worksheet[config.report_date_cell] = payload["report_date"]
    worksheet[config.institution_cell] = payload["form_metadata"]["institution"]
    if config.category_cell is not None:
        worksheet[config.category_cell] = _cost_statement_category_text(payload)

    current_row = config.data_start_row
    for row in payload["rows"]:
        worksheet[f"{config.index_column}{current_row}"] = row["index"]
        worksheet[f"{config.student_column}{current_row}"] = row["student_name"]
        worksheet[f"{config.group_column}{current_row}"] = row["group_name"]
        worksheet[f"{config.amount_column}{current_row}"] = row["total_amount"]
        current_row += 1

    _hide_rows(worksheet, current_row, config.total_row - 1)
    _clear_rows(worksheet, config.total_row, config.total_row)
    worksheet[f"{config.student_column}{config.total_row}"] = "Итого:"
    worksheet[f"{config.amount_column}{config.total_row}"] = payload["grand_total"]
    apply_prepared_by_binding(
        worksheet,
        prepared_by_short_name=payload["prepared_by_short_name"],
        binding=config.prepared_by_binding,
    )


def populate_meal_sheet_worksheet(
    worksheet: Worksheet,
    payload: dict,
    config: MealSheetTemplateConfig,
) -> None:
    _assert_capacity(len(payload["rows"]), config.capacity, "табеля")

    day_slots = _day_slot_columns(worksheet, config)
    _normalize_day_slot_widths(worksheet, day_slots, target_width=config.day_slot_width)
    if len(payload["days"]) > len(day_slots):
        raise ValueError("Для выбранного месяца в шаблоне не хватает столбцов по дням")

    worksheet[config.month_cell] = russian_month_name(payload["period_start"])
    worksheet[config.year_cell] = _year_label(payload["year"])
    worksheet[config.title_cell] = payload["title"]
    _protect_meal_sheet_period_header(worksheet, config)
    if config.student_header_cell is not None:
        worksheet[config.student_header_cell] = _meal_sheet_student_header_text(payload, config)
    if config.price_label_cell is not None:
        worksheet[config.price_label_cell] = _meal_price_label(payload["meal_type"])

    for slot_index, column_letter in enumerate(day_slots):
        day_payload = payload["days"][slot_index] if slot_index < len(payload["days"]) else None
        if day_payload is None:
            _clear_day_slot(worksheet, column_letter, config)
            worksheet.column_dimensions[column_letter].hidden = True
            continue

        worksheet.column_dimensions[column_letter].hidden = False
        worksheet[f"{column_letter}{config.header_day_row}"] = day_payload["day"]
        worksheet[f"{column_letter}{config.meal_label_row}"] = payload["meal_type_label"]
        worksheet[f"{column_letter}{config.price_row}"] = payload["day_prices"][day_payload["iso"]]

    _protect_meal_sheet_day_slots(worksheet, config, day_slots)
    _hide_non_day_columns_between_slots(worksheet, config, day_slots)

    current_row = config.data_start_row
    for row in payload["rows"]:
        worksheet[f"{config.index_column}{current_row}"] = row["index"]
        worksheet[f"{config.student_column}{current_row}"] = row["student_name"]
        worksheet[f"{config.group_column}{current_row}"] = row["group_name"]
        if config.date_column is not None:
            worksheet[f"{config.date_column}{current_row}"] = row["date_value"]

        for slot_index, column_letter in enumerate(day_slots):
            day_payload = payload["days"][slot_index] if slot_index < len(payload["days"]) else None
            worksheet[f"{column_letter}{current_row}"] = row["marks"].get(day_payload["iso"], "") if day_payload else None

        worksheet[f"{config.total_column}{current_row}"] = row["total_amount"]
        current_row += 1

    _hide_rows(worksheet, current_row, config.count_row - 1)
    _hide_empty_tail_columns(worksheet, config, day_slots, len(payload["days"]))
    _hide_trailing_columns_after_total(worksheet, config)
    worksheet[config.institution_cell] = payload["form_metadata"]["institution"]
    _clear_footer_borders(
        worksheet,
        start_row=config.amount_row + 1,
        end_row=config.signature_row - 1,
        start_column=day_slots[0] if day_slots else config.total_column,
        end_column=config.total_column,
    )

    for slot_index, column_letter in enumerate(day_slots):
        day_payload = payload["days"][slot_index] if slot_index < len(payload["days"]) else None
        if day_payload is None:
            worksheet[f"{column_letter}{config.count_row}"] = None
            worksheet[f"{column_letter}{config.amount_row}"] = None
            continue

        worksheet[f"{column_letter}{config.count_row}"] = payload["day_counts"][day_payload["iso"]]
        worksheet[f"{column_letter}{config.amount_row}"] = payload["day_amounts"][day_payload["iso"]]

    worksheet[f"{config.total_column}{config.count_row}"] = payload["total_count"]
    worksheet[f"{config.total_column}{config.amount_row}"] = payload["total_amount"]
    apply_prepared_by_binding(
        worksheet,
        prepared_by_short_name=payload["prepared_by_short_name"],
        binding=config.prepared_by_binding,
    )


def populate_cost_calculation_worksheet(
    worksheet: Worksheet,
    payload: dict,
    config: CostCalculationTemplateConfig,
    *,
    custom_values: dict[str, str] | None = None,
) -> None:
    _assert_capacity(len(payload["rows"]), config.capacity, "расчета")
    _apply_column_width_contract(worksheet, config.column_width_contract)

    daily_compensation_rate = _daily_compensation_rate(
        worksheet[config.daily_compensation_rate_cell].value,
        custom_values,
    )
    worksheet[config.category_line_cell] = payload["category_line"]
    worksheet[config.month_cell] = russian_month_name(payload["period_start"])
    worksheet[config.year_cell] = _year_label(payload["year"])
    worksheet[config.study_days_cell] = payload["study_day_count"]
    worksheet[config.daily_compensation_rate_cell] = daily_compensation_rate

    total_accrued_amount = 0.0
    total_meal_amount = 0.0
    total_payout_amount = 0.0

    current_row = config.data_start_row
    for row in payload["rows"]:
        accrued_amount = round(daily_compensation_rate * payload["study_day_count"], 2)
        payout_amount = round(accrued_amount - row["meal_amount"], 2)

        worksheet[f"{config.index_column}{current_row}"] = row["index"]
        worksheet[f"{config.student_column}{current_row}"] = row["student_name"]
        worksheet[f"{config.study_days_column}{current_row}"] = payload["study_day_count"]
        worksheet[f"{config.accrued_amount_column}{current_row}"] = accrued_amount
        worksheet[f"{config.meal_amount_column}{current_row}"] = row["meal_amount"]
        worksheet[f"{config.payout_amount_column}{current_row}"] = payout_amount
        worksheet[f"{config.note_column}{current_row}"] = row.get("note")

        total_accrued_amount += accrued_amount
        total_meal_amount += row["meal_amount"]
        total_payout_amount += payout_amount
        current_row += 1

    _hide_rows(worksheet, current_row, config.total_row - 1)
    worksheet[f"{config.accrued_amount_column}{config.total_row}"] = round(total_accrued_amount, 2)
    worksheet[f"{config.meal_amount_column}{config.total_row}"] = round(total_meal_amount, 2)
    worksheet[f"{config.payout_amount_column}{config.total_row}"] = round(total_payout_amount, 2)
    apply_prepared_by_binding(
        worksheet,
        prepared_by_short_name=payload["prepared_by_short_name"],
        binding=config.prepared_by_binding,
    )


def _day_slot_columns(worksheet: Worksheet, config: MealSheetTemplateConfig) -> list[str]:
    min_col, _, max_col, _ = range_boundaries(config.visible_range)
    columns: list[str] = []
    for column_index in range(min_col, max_col + 1):
        cell_value = worksheet.cell(config.header_day_row, column_index).value
        if isinstance(cell_value, bool):
            continue
        if isinstance(cell_value, (int, float)) and float(cell_value).is_integer() and 1 <= int(cell_value) <= 31:
            columns.append(get_column_letter(column_index))
    return columns


def _clear_day_slot(worksheet: Worksheet, column_letter: str, config: MealSheetTemplateConfig) -> None:
    worksheet[f"{column_letter}{config.header_day_row}"] = None
    worksheet[f"{column_letter}{config.meal_label_row}"] = None
    worksheet[f"{column_letter}{config.price_row}"] = None

    for row_index in range(config.data_start_row, config.count_row):
        worksheet[f"{column_letter}{row_index}"] = None

    worksheet[f"{column_letter}{config.count_row}"] = None
    worksheet[f"{column_letter}{config.amount_row}"] = None


def _normalize_day_slot_widths(
    worksheet: Worksheet,
    day_slots: list[str],
    *,
    target_width: float | None = None,
) -> None:
    if not day_slots:
        return

    if target_width is not None and target_width > 0:
        normalized_width = round(float(target_width), 6)
    else:
        widths = [float(worksheet.column_dimensions[column_letter].width or 0) for column_letter in day_slots]
        valid_widths = [width for width in widths if width > 0]
        if not valid_widths:
            return
        normalized_width = round(sum(valid_widths) / len(valid_widths), 6)

    for column_letter in day_slots:
        worksheet.column_dimensions[column_letter].width = normalized_width


def _protect_meal_sheet_period_header(worksheet: Worksheet, config: MealSheetTemplateConfig) -> None:
    month_cell = worksheet[config.month_cell]
    month_alignment = copy(month_cell.alignment)
    month_alignment.horizontal = month_alignment.horizontal or "center"
    month_alignment.vertical = month_alignment.vertical or "center"
    month_alignment.wrap_text = True
    month_alignment.shrink_to_fit = False
    month_cell.alignment = month_alignment

    for cell_ref in (config.year_cell, config.institution_cell):
        cell = worksheet[cell_ref]
        alignment = copy(cell.alignment)
        alignment.shrink_to_fit = False
        cell.alignment = alignment


def _protect_meal_sheet_day_slots(
    worksheet: Worksheet,
    config: MealSheetTemplateConfig,
    day_slots: list[str],
) -> None:
    for column_letter in day_slots:
        if worksheet.column_dimensions[column_letter].hidden:
            continue
        for row_index in range(config.header_day_row, config.amount_row + 1):
            cell = worksheet[f"{column_letter}{row_index}"]
            alignment = copy(cell.alignment)
            alignment.horizontal = alignment.horizontal or "center"
            alignment.vertical = alignment.vertical or "center"
            alignment.shrink_to_fit = False
            cell.alignment = alignment


def _merge_cost_statement_requisite_text_fields(worksheet: Worksheet) -> None:
    for range_string in ("C9:G9", "C15:G15"):
        if range_string not in {str(merged_range) for merged_range in worksheet.merged_cells.ranges}:
            worksheet.merge_cells(range_string)


def _apply_column_width_contract(worksheet: Worksheet, contract: ColumnWidthContract) -> None:
    if not contract:
        return

    for column_letter, width in contract:
        if width <= 0:
            continue
        worksheet.column_dimensions[column_letter].width = round(float(width), 6)


def _hide_rows(worksheet: Worksheet, start_row: int, end_row: int) -> None:
    if start_row > end_row:
        return
    for row_index in range(start_row, end_row + 1):
        worksheet.row_dimensions[row_index].hidden = True


def _clear_rows(worksheet: Worksheet, start_row: int, end_row: int) -> None:
    if start_row > end_row:
        return

    max_column = worksheet.max_column
    for row_index in range(start_row, end_row + 1):
        for column_index in range(1, max_column + 1):
            cell = worksheet.cell(row=row_index, column=column_index)
            if isinstance(cell, MergedCell):
                continue
            cell.value = None


def _clear_footer_borders(
    worksheet: Worksheet,
    *,
    start_row: int,
    end_row: int,
    start_column: str,
    end_column: str,
) -> None:
    if start_row > end_row:
        return

    start_index = column_index_from_string(start_column)
    end_index = column_index_from_string(end_column)
    for row_index in range(start_row, end_row + 1):
        for column_index in range(start_index, end_index + 1):
            worksheet.cell(row=row_index, column=column_index).border = Border()


def _hide_non_day_columns_between_slots(
    worksheet: Worksheet,
    config: MealSheetTemplateConfig,
    day_slots: list[str],
) -> None:
    if not day_slots:
        return

    slot_set = set(day_slots)
    for column_index in range(
        column_index_from_string(day_slots[0]),
        column_index_from_string(config.total_column),
    ):
        column_letter = get_column_letter(column_index)
        if column_letter in slot_set:
            continue
        worksheet.column_dimensions[column_letter].hidden = True


def _hide_empty_tail_columns(
    worksheet: Worksheet,
    config: MealSheetTemplateConfig,
    day_slots: list[str],
    populated_days: int,
) -> None:
    if not day_slots or populated_days <= 0:
        return

    last_used_day_column = day_slots[min(populated_days, len(day_slots)) - 1]
    for column_index in range(
        column_index_from_string(last_used_day_column) + 1,
        column_index_from_string(config.total_column),
    ):
        column_letter = get_column_letter(column_index)
        if _column_has_values(worksheet, column_letter, config.header_day_row, config.amount_row):
            continue
        worksheet.column_dimensions[column_letter].hidden = True


def _hide_trailing_columns_after_total(worksheet: Worksheet, config: MealSheetTemplateConfig) -> None:
    _, _, max_col, _ = range_boundaries(config.visible_range)
    total_column_index = column_index_from_string(config.total_column)
    for column_index in range(total_column_index + 1, max_col + 1):
        column_letter = get_column_letter(column_index)
        if _column_has_values(worksheet, column_letter, config.header_day_row, config.amount_row):
            continue
        worksheet.column_dimensions[column_letter].hidden = True


def _column_has_values(worksheet: Worksheet, column_letter: str, start_row: int, end_row: int) -> bool:
    for row_index in range(start_row, end_row + 1):
        if worksheet[f"{column_letter}{row_index}"].value not in {None, ""}:
            return True
    return False


def _assert_capacity(actual_rows: int, capacity: int, document_label: str) -> None:
    if actual_rows <= capacity:
        return
    raise ValueError(f"Для выбранного {document_label} недостаточно строк в эталонном шаблоне")


def _year_label(year: int) -> str:
    return f"{year} г."


def _cost_statement_category_text(payload: dict) -> str:
    category_label = payload.get("category_label")
    if payload.get("category_code") == "all" or not category_label:
        return 'категория "все категории"'
    return f'категория "{category_label}"'


def _meal_sheet_student_header_text(payload: dict, config: MealSheetTemplateConfig) -> str:
    if config.date_column is None:
        return "Фамилия Имя Отчество студента, группа"

    category_label = payload.get("category_label")
    if payload.get("category_code") == "all" or not category_label:
        return "Фамилия Имя Отчество студента, группа, дата действия подтверждающего документа"

    return f'Фамилия Имя Отчество студента, группа, дата действия справки о категории "{category_label}"'


def _numeric_cell_value(value) -> float:
    if isinstance(value, (int, float)):
        return float(value)
    raise ValueError("В шаблоне расчета стоимости питания отсутствует корректный дневной норматив")


def _daily_compensation_rate(template_value, custom_values: dict[str, str] | None) -> float:
    raw_value = (custom_values or {}).get("dailyCompensationRate")
    if raw_value is not None and str(raw_value).strip():
        return parse_decimal_metadata_value(raw_value, field_label="Дневной норматив компенсации")

    return _numeric_cell_value(template_value)


def _meal_price_label(meal_type: str) -> str:
    if meal_type == "breakfast":
        return "Стоимость завтрака:"
    return "Стоимость обеда:"
