from __future__ import annotations

from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import range_boundaries

MM_PER_INCH = 25.4
DEFAULT_PAPER_SIZE_MM = (210.0, 297.0)
PAPER_SIZES_MM: dict[int, tuple[float, float]] = {
    1: (215.9, 279.4),
    5: (215.9, 355.6),
    8: (297.0, 420.0),
    9: (210.0, 297.0),
}


def merged_maps(
    sheet,
    visible_rows: list[int],
    visible_columns: list[int],
) -> tuple[dict[tuple[int, int], tuple[int, int]], set[tuple[int, int]]]:
    visible_row_set = set(visible_rows)
    visible_column_set = set(visible_columns)
    merged_tops: dict[tuple[int, int], tuple[int, int]] = {}
    merged_children: set[tuple[int, int]] = set()

    for merged_range in sheet.merged_cells.ranges:
        min_col, min_row, max_col, max_row = range_boundaries(str(merged_range))
        rowspan = sum(1 for row_index in range(min_row, max_row + 1) if row_index in visible_row_set)
        colspan = sum(1 for column_index in range(min_col, max_col + 1) if column_index in visible_column_set)
        if rowspan == 0 or colspan == 0:
            continue

        top = (min_row, min_col)
        merged_tops[top] = (rowspan, colspan)
        for row_index in range(min_row, max_row + 1):
            for column_index in range(min_col, max_col + 1):
                if (row_index, column_index) == top:
                    continue
                if row_index in visible_row_set and column_index in visible_column_set:
                    merged_children.add((row_index, column_index))

    return merged_tops, merged_children


def overflow_visible_cells(
    sheet,
    visible_rows: list[int],
    visible_columns: list[int],
    merged_tops: dict[tuple[int, int], tuple[int, int]],
    merged_children: set[tuple[int, int]],
) -> dict[tuple[int, int], str]:
    visible_column_positions = {column_index: position for position, column_index in enumerate(visible_columns)}
    overflow_cells: dict[tuple[int, int], str] = {}

    for row_index in visible_rows:
        for column_index in visible_columns:
            coordinate = (row_index, column_index)
            if coordinate in merged_tops or coordinate in merged_children:
                continue

            cell = sheet.cell(row=row_index, column=column_index)
            if not cell_can_overflow(cell):
                continue

            start_position = visible_column_positions[column_index]
            for direction in overflow_directions(cell):
                if has_blank_overflow_run(
                    sheet,
                    row_index=row_index,
                    visible_columns=visible_columns,
                    start_position=start_position,
                    merged_tops=merged_tops,
                    merged_children=merged_children,
                    direction=direction,
                ):
                    overflow_cells[coordinate] = direction
                    break

    return overflow_cells


def cell_can_overflow(cell) -> bool:
    if not isinstance(cell.value, str):
        return False
    if not cell.value.strip():
        return False

    alignment = cell.alignment
    if alignment and alignment.wrap_text:
        return False
    if alignment and alignment.textRotation not in {None, 0}:
        return False
    if _cell_has_visible_border(cell):
        return False
    if _cell_has_solid_fill(cell):
        return False

    return True


def overflow_directions(cell) -> tuple[str, ...]:
    alignment = cell.alignment
    horizontal = alignment.horizontal if alignment else None
    if horizontal == "right":
        return ("left",)
    if horizontal in {"center", "centerContinuous"}:
        return ("left", "right")
    return ("right",)


def has_blank_overflow_run(
    sheet,
    *,
    row_index: int,
    visible_columns: list[int],
    start_position: int,
    merged_tops: dict[tuple[int, int], tuple[int, int]],
    merged_children: set[tuple[int, int]],
    direction: str,
) -> bool:
    step = -1 if direction == "left" else 1
    position = start_position + step
    found_blank = False

    while 0 <= position < len(visible_columns):
        column_index = visible_columns[position]
        coordinate = (row_index, column_index)
        if coordinate in merged_tops or coordinate in merged_children:
            break

        neighbor = sheet.cell(row=row_index, column=column_index)
        if neighbor.value not in {None, ""}:
            break

        found_blank = True
        position += step

    return found_blank


def row_height_mm(sheet, row_index: int) -> float:
    dimension = sheet.row_dimensions[row_index]
    height = dimension.height or sheet.sheet_format.defaultRowHeight or 15
    return height * MM_PER_INCH / 72


def worksheet_widths(
    sheet,
    visible_columns: list[int],
) -> tuple[dict[int, float], dict[int, float], float, float, float]:
    screen_column_widths = {index: _column_width_mm(sheet, index) for index in visible_columns}
    screen_width = sum(screen_column_widths.values())
    printable_width = _printable_width_mm(sheet)
    print_width = _target_print_width_mm(sheet, screen_width, printable_width)

    if screen_width <= 0:
        return screen_column_widths, screen_column_widths, 1.0, printable_width, printable_width
    if abs(print_width - screen_width) < 0.01:
        return screen_column_widths, screen_column_widths, screen_width, print_width, printable_width

    ratio = print_width / screen_width
    print_column_widths = {index: width * ratio for index, width in screen_column_widths.items()}
    return screen_column_widths, print_column_widths, screen_width, print_width, printable_width


def _cell_has_visible_border(cell) -> bool:
    border = cell.border
    if border is None:
        return False

    return any(
        getattr(side, "style", None) is not None
        for side in (border.top, border.right, border.bottom, border.left)
    )


def _cell_has_solid_fill(cell) -> bool:
    fill = cell.fill
    return bool(fill and fill.fill_type == "solid")


def _column_width_mm(sheet, column_index: int) -> float:
    dimension = sheet.column_dimensions[get_column_letter(column_index)]
    width = dimension.width or 8.43
    if width < 1:
        pixels = width * 12
    else:
        pixels = width * 7 + 5
    return pixels * MM_PER_INCH / 96


def _target_print_width_mm(sheet, raw_total_width: float, printable_width: float) -> float:
    if raw_total_width <= 0:
        return printable_width

    scale_ratio = _worksheet_scale_ratio(sheet)
    if scale_ratio is not None:
        scaled_width = raw_total_width * scale_ratio
    elif _worksheet_uses_fit_to_width(sheet):
        scaled_width = printable_width
    else:
        scaled_width = raw_total_width

    return max(min(scaled_width, printable_width), 1.0)


def _printable_width_mm(sheet) -> float:
    page_width_mm, _ = _page_size_mm(sheet.page_setup.paperSize, sheet.page_setup.orientation)
    margins = sheet.page_margins
    left_margin_mm = (margins.left or 0.7) * MM_PER_INCH
    right_margin_mm = (margins.right or 0.7) * MM_PER_INCH
    return max(page_width_mm - left_margin_mm - right_margin_mm, 1.0)


def _page_size_mm(paper_size, orientation: str | None) -> tuple[float, float]:
    width_mm, height_mm = _paper_dimensions_mm(paper_size)
    if orientation == "landscape":
        return max(width_mm, height_mm), min(width_mm, height_mm)
    return min(width_mm, height_mm), max(width_mm, height_mm)


def _paper_dimensions_mm(paper_size) -> tuple[float, float]:
    try:
        return PAPER_SIZES_MM[int(paper_size)]
    except (KeyError, TypeError, ValueError):
        return DEFAULT_PAPER_SIZE_MM


def _worksheet_scale_ratio(sheet) -> float | None:
    scale = sheet.page_setup.scale
    if scale in {None, ""}:
        return None

    try:
        numeric_scale = float(scale)
    except (TypeError, ValueError):
        return None

    if numeric_scale <= 0:
        return None
    return numeric_scale / 100.0


def _worksheet_uses_fit_to_width(sheet) -> bool:
    fit_to_width = sheet.page_setup.fitToWidth
    if fit_to_width in {None, ""}:
        return False

    try:
        return int(fit_to_width) > 0
    except (TypeError, ValueError):
        return False
