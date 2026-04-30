from __future__ import annotations

import math
import os
import re
import shutil
import socket
import subprocess
import sys
import tempfile
import time
from dataclasses import dataclass
from importlib import import_module
from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries
from openpyxl.worksheet.properties import PageSetupProperties

from .worksheet_layout import printable_width_mm, worksheet_print_area_ranges, worksheet_print_area_width_mm

CALC_PDF_EXPORT_FILTER = "pdf:calc_pdf_Export"
DEFAULT_MAX_AUTO_SCALE_PERCENT = 125
MM_100_PER_MM = 100
A4_WIDTH_MM = 210
A4_HEIGHT_MM = 297
DEFAULT_PDF_MARGIN_MM = 5


class XlsxToPdfConversionError(RuntimeError):
    """Controlled failure for XLSX to PDF conversion."""


@dataclass(frozen=True, slots=True)
class PdfScaleDiagnostic:
    sheet_title: str
    print_area: tuple[str, ...]
    print_area_width_mm: float
    printable_width_mm: float
    target_scale_percent: int
    applied: bool


@dataclass(frozen=True, slots=True)
class PdfPageGeometry:
    media_width: float
    media_height: float
    crop_width: float
    crop_height: float
    rotate: int

    @property
    def is_landscape(self) -> bool:
        return self.crop_width > self.crop_height


def convert_xlsx_to_pdf(
    workbook_bytes: bytes,
    *,
    libreoffice_bin: str,
    timeout_seconds: int,
    max_auto_scale_percent: int = DEFAULT_MAX_AUTO_SCALE_PERCENT,
) -> bytes:
    with tempfile.TemporaryDirectory(prefix="accounting-pdf-") as tmp_dir_name:
        tmp_dir = Path(tmp_dir_name)
        xlsx_path = tmp_dir / "document.xlsx"
        pdf_path = tmp_dir / "document.pdf"
        profile_dir = tmp_dir / "libreoffice-profile"
        profile_dir.mkdir(parents=True, exist_ok=True)
        pdf_workbook_bytes, _ = prepare_xlsx_for_pdf(
            workbook_bytes,
            max_auto_scale_percent=max_auto_scale_percent,
        )
        xlsx_path.write_bytes(pdf_workbook_bytes)
        _convert_xlsx_to_pdf_with_uno(
            xlsx_path=xlsx_path,
            pdf_path=pdf_path,
            profile_dir=profile_dir,
            libreoffice_bin=libreoffice_bin,
            timeout_seconds=timeout_seconds,
        )
        if not pdf_path.exists():
            raise XlsxToPdfConversionError("LibreOffice did not produce a PDF file")

        pdf_bytes = pdf_path.read_bytes()
        if not pdf_bytes.startswith(b"%PDF"):
            raise XlsxToPdfConversionError("LibreOffice produced an invalid PDF file")
        validate_pdf_page_geometry(
            pdf_bytes,
            expected_landscape=_workbook_expects_landscape(pdf_workbook_bytes),
        )
        return pdf_bytes


def _convert_xlsx_to_pdf_with_uno(
    *,
    xlsx_path: Path,
    pdf_path: Path,
    profile_dir: Path,
    libreoffice_bin: str,
    timeout_seconds: int,
) -> None:
    port = _free_local_port()
    command = _libreoffice_uno_command(libreoffice_bin, profile_dir=profile_dir, port=port)
    try:
        process = subprocess.Popen(
            command,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
        )
    except FileNotFoundError as exc:
        raise XlsxToPdfConversionError("LibreOffice executable was not found") from exc
    except OSError as exc:
        raise XlsxToPdfConversionError("LibreOffice could not be started") from exc

    document = None
    try:
        uno_module = _uno_module(libreoffice_bin)
        context = _connect_to_libreoffice_uno(uno_module, port=port, process=process, timeout_seconds=timeout_seconds)
        service_manager = context.ServiceManager
        desktop = service_manager.createInstanceWithContext("com.sun.star.frame.Desktop", context)
        document = desktop.loadComponentFromURL(
            uno_module.systemPathToFileUrl(str(xlsx_path)),
            "_blank",
            0,
            (
                _uno_property(uno_module, "Hidden", True),
                _uno_property(uno_module, "ReadOnly", False),
            ),
        )
        if document is None:
            raise XlsxToPdfConversionError("LibreOffice could not open generated XLSX")

        source_workbook = load_workbook(xlsx_path, data_only=False)
        _apply_calc_page_styles_for_pdf(uno_module, document, source_workbook)
        document.storeToURL(
            uno_module.systemPathToFileUrl(str(pdf_path)),
            (
                _uno_property(uno_module, "FilterName", "calc_pdf_Export"),
            ),
        )
    except XlsxToPdfConversionError:
        raise
    except Exception as exc:
        raise XlsxToPdfConversionError("LibreOffice UNO PDF export failed") from exc
    finally:
        if document is not None:
            try:
                document.close(True)
            except Exception:
                try:
                    document.dispose()
                except Exception:
                    pass
        _stop_libreoffice_process(process)


def _libreoffice_uno_command(libreoffice_bin: str, *, profile_dir: Path, port: int) -> list[str]:
    return [
        libreoffice_bin,
        "--headless",
        "--invisible",
        "--nologo",
        "--nodefault",
        "--norestore",
        "--nofirststartwizard",
        "--nolockcheck",
        f"-env:UserInstallation={profile_dir.as_uri()}",
        f"--accept=socket,host=127.0.0.1,port={port};urp;StarOffice.ComponentContext",
    ]


def prepare_xlsx_for_pdf(
    workbook_bytes: bytes,
    *,
    max_auto_scale_percent: int = DEFAULT_MAX_AUTO_SCALE_PERCENT,
) -> tuple[bytes, list[PdfScaleDiagnostic]]:
    max_scale_percent = max(int(max_auto_scale_percent or 100), 100)
    workbook = load_workbook(BytesIO(workbook_bytes), data_only=False)

    diagnostics: list[PdfScaleDiagnostic] = []
    changed = False
    for worksheet in workbook.worksheets:
        diagnostic = _apply_pdf_auto_scale(worksheet, max_scale_percent=max_scale_percent)
        diagnostics.append(diagnostic)
        changed = changed or diagnostic.applied

    if not changed:
        return workbook_bytes, diagnostics

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer.getvalue(), diagnostics


def validate_pdf_page_geometry(pdf_bytes: bytes, *, expected_landscape: bool) -> PdfPageGeometry:
    geometry = inspect_pdf_first_page_geometry(pdf_bytes)
    if expected_landscape and (not geometry.is_landscape or geometry.rotate != 0):
        raise XlsxToPdfConversionError("LibreOffice produced a rotated or portrait PDF page")
    return geometry


def inspect_pdf_first_page_geometry(pdf_bytes: bytes) -> PdfPageGeometry:
    media_box = _extract_pdf_box(pdf_bytes, b"MediaBox")
    crop_box = _extract_pdf_box(pdf_bytes, b"CropBox") or media_box
    if media_box is None or crop_box is None:
        raise XlsxToPdfConversionError("Could not inspect PDF page geometry")

    rotate_match = re.search(rb"/Rotate\s+(-?\d+)", pdf_bytes)
    rotate = int(rotate_match.group(1)) if rotate_match else 0
    return PdfPageGeometry(
        media_width=abs(media_box[2] - media_box[0]),
        media_height=abs(media_box[3] - media_box[1]),
        crop_width=abs(crop_box[2] - crop_box[0]),
        crop_height=abs(crop_box[3] - crop_box[1]),
        rotate=rotate,
    )


def _apply_pdf_auto_scale(worksheet, *, max_scale_percent: int) -> PdfScaleDiagnostic:
    print_area = worksheet_print_area_ranges(worksheet)
    print_area_width = worksheet_print_area_width_mm(worksheet)
    printable_width = printable_width_mm(worksheet, orientation="landscape")
    target_scale_percent = _target_pdf_scale_percent(
        print_area_width_mm=print_area_width,
        printable_width_mm=printable_width,
        max_scale_percent=max_scale_percent,
    )

    if target_scale_percent <= 100:
        return PdfScaleDiagnostic(
            sheet_title=worksheet.title,
            print_area=print_area,
            print_area_width_mm=print_area_width,
            printable_width_mm=printable_width,
            target_scale_percent=100,
            applied=False,
        )

    if worksheet.sheet_properties.pageSetUpPr is None:
        worksheet.sheet_properties.pageSetUpPr = PageSetupProperties()
    worksheet.sheet_properties.pageSetUpPr.fitToPage = False
    worksheet.page_setup.fitToWidth = 0
    worksheet.page_setup.fitToHeight = 0
    worksheet.page_setup.scale = target_scale_percent

    return PdfScaleDiagnostic(
        sheet_title=worksheet.title,
        print_area=print_area,
        print_area_width_mm=print_area_width,
        printable_width_mm=printable_width,
        target_scale_percent=target_scale_percent,
        applied=True,
    )


def _target_pdf_scale_percent(
    *,
    print_area_width_mm: float,
    printable_width_mm: float,
    max_scale_percent: int,
) -> int:
    if print_area_width_mm <= 0 or printable_width_mm <= 0:
        return 100

    available_scale = printable_width_mm / print_area_width_mm
    target_scale = min(available_scale, max_scale_percent / 100)
    target_scale = max(target_scale, 1.0)
    return max(math.floor(target_scale * 100), 100)


def _apply_calc_page_styles_for_pdf(uno_module, document, source_workbook) -> None:
    page_styles = document.getStyleFamilies().getByName("PageStyles")
    spreadsheets = document.getSheets()
    for sheet_index, source_sheet in enumerate(source_workbook.worksheets):
        spreadsheet = spreadsheets.getByIndex(sheet_index)
        page_style = page_styles.getByName(spreadsheet.PageStyle)
        is_landscape = source_sheet.page_setup.orientation == "landscape"
        _configure_calc_page_style(
            uno_module,
            page_style,
            is_landscape=is_landscape,
            fit_to_width=int(source_sheet.page_setup.fitToWidth or 0),
            fit_to_height=int(source_sheet.page_setup.fitToHeight or 0),
            page_scale=_page_scale(source_sheet),
        )
        _set_calc_print_areas(uno_module, spreadsheet, sheet_index=sheet_index, print_areas=worksheet_print_area_ranges(source_sheet))


def _configure_calc_page_style(
    uno_module,
    page_style,
    *,
    is_landscape: bool,
    fit_to_width: int,
    fit_to_height: int,
    page_scale: int | None,
) -> None:
    page_style.IsLandscape = is_landscape
    page_style.Width = _mm100(A4_HEIGHT_MM if is_landscape else A4_WIDTH_MM)
    page_style.Height = _mm100(A4_WIDTH_MM if is_landscape else A4_HEIGHT_MM)
    for property_name in ("LeftMargin", "RightMargin", "TopMargin", "BottomMargin"):
        setattr(page_style, property_name, _mm100(DEFAULT_PDF_MARGIN_MM))

    try:
        page_style.PaperFormat = uno_module.getConstantByName("com.sun.star.view.PaperFormat.A4")
    except Exception:
        pass

    if page_scale is not None and page_scale > 100:
        page_style.ScaleToPagesX = 0
        page_style.ScaleToPagesY = 0
        page_style.PageScale = page_scale
        return

    page_style.ScaleToPagesX = fit_to_width
    page_style.ScaleToPagesY = fit_to_height
    page_style.PageScale = 100


def _set_calc_print_areas(uno_module, spreadsheet, *, sheet_index: int, print_areas: tuple[str, ...]) -> None:
    if not print_areas or not hasattr(spreadsheet, "setPrintAreas"):
        return

    addresses = []
    for range_string in print_areas:
        min_col, min_row, max_col, max_row = range_boundaries(range_string)
        address = uno_module.createUnoStruct("com.sun.star.table.CellRangeAddress")
        address.Sheet = sheet_index
        address.StartColumn = min_col - 1
        address.StartRow = min_row - 1
        address.EndColumn = max_col - 1
        address.EndRow = max_row - 1
        addresses.append(address)
    spreadsheet.setPrintAreas(tuple(addresses))


def _page_scale(worksheet) -> int | None:
    scale = worksheet.page_setup.scale
    if scale in {None, ""}:
        return None
    try:
        return int(scale)
    except (TypeError, ValueError):
        return None


def _workbook_expects_landscape(workbook_bytes: bytes) -> bool:
    workbook = load_workbook(BytesIO(workbook_bytes), read_only=False, data_only=False)
    try:
        return any(worksheet.page_setup.orientation == "landscape" for worksheet in workbook.worksheets)
    finally:
        workbook.close()


def _connect_to_libreoffice_uno(uno_module, *, port: int, process: subprocess.Popen, timeout_seconds: int):
    deadline = time.monotonic() + max(timeout_seconds, 1)
    last_error: Exception | None = None
    while time.monotonic() < deadline:
        if process.poll() is not None:
            raise XlsxToPdfConversionError("LibreOffice exited before UNO connection was ready")
        try:
            local_context = uno_module.getComponentContext()
            resolver = local_context.ServiceManager.createInstanceWithContext(
                "com.sun.star.bridge.UnoUrlResolver",
                local_context,
            )
            return resolver.resolve(
                f"uno:socket,host=127.0.0.1,port={port};urp;StarOffice.ComponentContext"
            )
        except Exception as exc:
            last_error = exc
            time.sleep(0.2)

    raise XlsxToPdfConversionError("LibreOffice UNO connection timed out") from last_error


def _free_local_port() -> int:
    with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as server:
        server.bind(("127.0.0.1", 0))
        return int(server.getsockname()[1])


def _uno_module(libreoffice_bin: str):
    _add_libreoffice_program_dir_to_python_path(libreoffice_bin)
    try:
        return import_module("uno")
    except ImportError as exc:
        raise XlsxToPdfConversionError("LibreOffice UNO Python bridge is not available") from exc


def _add_libreoffice_program_dir_to_python_path(libreoffice_bin: str) -> None:
    executable = Path(libreoffice_bin)
    if not executable.is_absolute():
        resolved = shutil.which(libreoffice_bin)
        if resolved is None:
            return
        executable = Path(resolved)

    program_dir = executable.resolve().parent
    program_dir_string = str(program_dir)
    if program_dir_string not in sys.path:
        sys.path.insert(0, program_dir_string)
    os.environ["PATH"] = f"{program_dir_string}{os.pathsep}{os.environ.get('PATH', '')}"
    if hasattr(os, "add_dll_directory"):
        try:
            os.add_dll_directory(program_dir_string)
        except OSError:
            pass


def _uno_property(uno_module, name: str, value):
    prop = uno_module.createUnoStruct("com.sun.star.beans.PropertyValue")
    prop.Name = name
    prop.Value = value
    return prop


def _stop_libreoffice_process(process: subprocess.Popen) -> None:
    if process.poll() is not None:
        return
    process.terminate()
    try:
        process.wait(timeout=5)
    except subprocess.TimeoutExpired:
        process.kill()
        process.wait(timeout=5)


def _extract_pdf_box(pdf_bytes: bytes, box_name: bytes) -> tuple[float, float, float, float] | None:
    number = rb"[-+]?\d+(?:\.\d+)?"
    match = re.search(
        rb"/" + re.escape(box_name) + rb"\s*\[\s*(" + number + rb")\s+(" + number + rb")\s+(" + number + rb")\s+(" + number + rb")\s*\]",
        pdf_bytes,
    )
    if match is None:
        return None
    return tuple(float(part) for part in match.groups())


def _mm100(value_mm: int | float) -> int:
    return int(round(float(value_mm) * MM_100_PER_MM))
