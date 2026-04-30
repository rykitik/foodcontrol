from __future__ import annotations

from dataclasses import dataclass
from typing import Literal

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import column_index_from_string, get_column_letter

from app.utils.report_generator import russian_month_name
from .template_config import PreparedByBinding

TITLE_ROW = 7
SPACER_ROW = 8
DAY_COLUMN_START = 5
MAX_DAY_COUNT = 31
DAY_COLUMN_WIDTH = 3.0
BREAKFAST_TOTAL_COLUMN = "AJ"
LUNCH_TOTAL_COLUMN = "AK"
AMOUNT_TOTAL_COLUMN = "AL"
AMOUNT_TOTAL_COLUMN_INDEX = column_index_from_string(AMOUNT_TOTAL_COLUMN)
TITLE_CELL = f"A{TITLE_ROW}"
TITLE_MERGE_RANGE = f"A{TITLE_ROW}:P{TITLE_ROW}"
PERIOD_PREFIX_CELL = f"Q{TITLE_ROW}"
PERIOD_PREFIX_MERGE_RANGE = f"Q{TITLE_ROW}:R{TITLE_ROW}"
MONTH_CELL = f"S{TITLE_ROW}"
MONTH_MERGE_RANGE = f"S{TITLE_ROW}:X{TITLE_ROW}"
YEAR_CELL = f"AB{TITLE_ROW}"
YEAR_MERGE_RANGE = f"AB{TITLE_ROW}:AF{TITLE_ROW}"
INSTITUTION_CELL = f"A{SPACER_ROW}"
INSTITUTION_MERGE_RANGE = f"A{SPACER_ROW}:{AMOUNT_TOTAL_COLUMN}{SPACER_ROW}"
HEADER_ROW = 9
DAY_MARK_ROW = 10
PRICE_ROW = 11
DATA_START_ROW = 12


@dataclass(frozen=True, slots=True)
class CombinedMealSheetConfig:
    visible_range: str
    prepared_by_binding: PreparedByBinding
    institution_cell: str = INSTITUTION_CELL
    document_type: Literal["combined_meal_sheet"] = "combined_meal_sheet"
    page_orientation: Literal["landscape"] = "landscape"


def build_combined_meal_sheet_workbook(payload: dict) -> tuple[Workbook, CombinedMealSheetConfig]:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = _safe_sheet_title(f"табель общий {payload['category_name']}")
    worksheet.sheet_view.showGridLines = False

    row_count = max(len(payload["rows"]), 1)
    summary_start_row = DATA_START_ROW + row_count
    signature_row = summary_start_row + 4
    config = CombinedMealSheetConfig(
        visible_range=f"A{TITLE_ROW}:{AMOUNT_TOTAL_COLUMN}{signature_row}",
        prepared_by_binding=PreparedByBinding(
            cell=f"A{signature_row}",
            mode="prefixed_text_cell",
            prefix="Бухгалтер ___________________",
        ),
    )

    _setup_page(worksheet, config)
    _setup_columns(worksheet, payload)
    _write_header(worksheet, payload, config)
    _write_rows(worksheet, payload, row_count)
    _write_totals(worksheet, payload, summary_start_row)
    _write_signature(worksheet, payload, config)
    _style_document(worksheet, signature_row)
    return workbook, config


def _setup_page(worksheet, config: CombinedMealSheetConfig) -> None:
    worksheet.print_area = config.visible_range
    worksheet.page_setup.orientation = "landscape"
    worksheet.page_setup.paperSize = "9"
    worksheet.page_setup.fitToWidth = 1
    worksheet.page_setup.fitToHeight = 0
    worksheet.sheet_properties.pageSetUpPr.fitToPage = True
    worksheet.page_margins.left = 0.25
    worksheet.page_margins.right = 0.25
    worksheet.page_margins.top = 0.35
    worksheet.page_margins.bottom = 0.35
    worksheet.page_margins.header = 0.1
    worksheet.page_margins.footer = 0.1
    worksheet.print_title_rows = f"{TITLE_ROW}:{PRICE_ROW}"
    worksheet.freeze_panes = f"{get_column_letter(DAY_COLUMN_START)}{DATA_START_ROW}"


def _setup_columns(worksheet, payload: dict) -> None:
    widths = {
        "A": 4.0,
        "B": 23.0,
        "C": 8.5,
        "D": 12.0,
        BREAKFAST_TOTAL_COLUMN: 6.5,
        LUNCH_TOTAL_COLUMN: 6.5,
        AMOUNT_TOTAL_COLUMN: 9.0,
    }
    for column_letter, width in widths.items():
        worksheet.column_dimensions[column_letter].width = width

    active_day_count = len(payload["days"])
    for offset in range(MAX_DAY_COUNT):
        column_letter = get_column_letter(DAY_COLUMN_START + offset)
        worksheet.column_dimensions[column_letter].width = DAY_COLUMN_WIDTH
        worksheet.column_dimensions[column_letter].hidden = offset >= active_day_count


def _write_header(worksheet, payload: dict, config: CombinedMealSheetConfig) -> None:
    worksheet.merge_cells(TITLE_MERGE_RANGE)
    worksheet.merge_cells(PERIOD_PREFIX_MERGE_RANGE)
    worksheet.merge_cells(MONTH_MERGE_RANGE)
    worksheet.merge_cells(YEAR_MERGE_RANGE)
    worksheet.merge_cells(INSTITUTION_MERGE_RANGE)
    for column_letter in ("A", "B", "C", "D"):
        worksheet.merge_cells(f"{column_letter}{HEADER_ROW}:{column_letter}{PRICE_ROW}")

    worksheet[TITLE_CELL] = payload["title"]
    worksheet[PERIOD_PREFIX_CELL] = "за "
    worksheet[MONTH_CELL] = russian_month_name(payload["period_start"])
    worksheet[YEAR_CELL] = f"{payload['year']} г."
    worksheet[config.institution_cell] = payload["form_metadata"]["institution"]

    fixed_headers = {
        "A": "№",
        "B": "Фамилия, имя, отчество",
        "C": "Группа",
        "D": "Категория",
        BREAKFAST_TOTAL_COLUMN: "Завтрак",
        LUNCH_TOTAL_COLUMN: "Обед",
        AMOUNT_TOTAL_COLUMN: "Сумма",
    }
    for column_letter, value in fixed_headers.items():
        worksheet[f"{column_letter}{HEADER_ROW}"] = value
        if column_letter not in {"A", "B", "C", "D"}:
            worksheet[f"{column_letter}{DAY_MARK_ROW}"] = "итого"
            worksheet[f"{column_letter}{PRICE_ROW}"] = ""

    for offset, day in enumerate(payload["days"]):
        column_letter = get_column_letter(DAY_COLUMN_START + offset)
        worksheet[f"{column_letter}{HEADER_ROW}"] = day["day"]
        worksheet[f"{column_letter}{DAY_MARK_ROW}"] = "З/О"
        worksheet[f"{column_letter}{PRICE_ROW}"] = ""


def _write_rows(worksheet, payload: dict, row_count: int) -> None:
    rows = payload["rows"]
    for offset in range(row_count):
        row_index = DATA_START_ROW + offset
        row = rows[offset] if offset < len(rows) else None
        if row is None:
            continue

        worksheet[f"A{row_index}"] = row["index"]
        worksheet[f"B{row_index}"] = row["student_name"]
        worksheet[f"C{row_index}"] = row["group_name"]
        worksheet[f"D{row_index}"] = row["category_name"]
        worksheet[f"{BREAKFAST_TOTAL_COLUMN}{row_index}"] = row["breakfast_count"]
        worksheet[f"{LUNCH_TOTAL_COLUMN}{row_index}"] = row["lunch_count"]
        worksheet[f"{AMOUNT_TOTAL_COLUMN}{row_index}"] = row["total_amount"]

        for day_offset, day in enumerate(payload["days"]):
            column_letter = get_column_letter(DAY_COLUMN_START + day_offset)
            worksheet[f"{column_letter}{row_index}"] = row["marks"].get(day["iso"], "")


def _write_totals(worksheet, payload: dict, summary_start_row: int) -> None:
    total_rows = (
        (
            "ИТОГО завтраков",
            "breakfast",
            payload["meal_totals"]["breakfast"]["count"],
            None,
            payload["meal_totals"]["breakfast"]["amount"],
        ),
        (
            "ИТОГО обедов",
            "lunch",
            None,
            payload["meal_totals"]["lunch"]["count"],
            payload["meal_totals"]["lunch"]["amount"],
        ),
        (
            "ИТОГО всего",
            "total",
            payload["meal_totals"]["breakfast"]["count"],
            payload["meal_totals"]["lunch"]["count"],
            payload["total_amount"],
        ),
    )

    for offset, (label, meal_key, breakfast_total, lunch_total, amount_total) in enumerate(total_rows):
        row_index = summary_start_row + offset
        worksheet[f"B{row_index}"] = label
        worksheet[f"{BREAKFAST_TOTAL_COLUMN}{row_index}"] = breakfast_total
        worksheet[f"{LUNCH_TOTAL_COLUMN}{row_index}"] = lunch_total
        worksheet[f"{AMOUNT_TOTAL_COLUMN}{row_index}"] = amount_total
        for day_offset, day in enumerate(payload["days"]):
            column_letter = get_column_letter(DAY_COLUMN_START + day_offset)
            if meal_key == "total":
                value = payload["day_totals"][day["iso"]]["count"]
            else:
                value = payload["day_totals"][day["iso"]][meal_key]["count"]
            worksheet[f"{column_letter}{row_index}"] = value or ""


def _write_signature(worksheet, payload: dict, config: CombinedMealSheetConfig) -> None:
    worksheet.merge_cells(f"A{int(config.prepared_by_binding.cell[1:])}:{AMOUNT_TOTAL_COLUMN}{int(config.prepared_by_binding.cell[1:])}")
    worksheet[config.prepared_by_binding.cell] = (
        f"{config.prepared_by_binding.prefix}{payload['prepared_by_short_name']}"
    )


def _style_document(worksheet, signature_row: int) -> None:
    thin = Side(style="thin", color="000000")
    border = Border(top=thin, right=thin, bottom=thin, left=thin)
    header_font = Font(name="Arial", size=14, bold=True)
    for cell_ref in (TITLE_CELL, PERIOD_PREFIX_CELL, MONTH_CELL, YEAR_CELL):
        worksheet[cell_ref].font = header_font
    worksheet[MONTH_CELL].font = Font(name="Arial", size=14, bold=True, color="0033CC")
    worksheet[TITLE_CELL].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    worksheet[MONTH_CELL].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True, shrink_to_fit=True)
    worksheet[MONTH_CELL].border = Border(bottom=thin)
    worksheet[INSTITUTION_CELL].font = Font(name="Arial", size=12, bold=True)
    worksheet[INSTITUTION_CELL].alignment = Alignment(horizontal="right", vertical="center")
    worksheet.row_dimensions[TITLE_ROW].height = 32.25
    worksheet.row_dimensions[SPACER_ROW].height = 16.5

    for row in worksheet.iter_rows(
        min_row=HEADER_ROW,
        max_row=signature_row - 2,
        min_col=1,
        max_col=AMOUNT_TOTAL_COLUMN_INDEX,
    ):
        for cell in row:
            cell.font = Font(name="Arial", size=8)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border

    for row_index in (HEADER_ROW, DAY_MARK_ROW, PRICE_ROW):
        for cell in worksheet[row_index]:
            cell.font = Font(name="Arial", size=8, bold=True)
        worksheet.row_dimensions[row_index].height = {
            HEADER_ROW: 21.75,
            DAY_MARK_ROW: 14.25,
            PRICE_ROW: 39,
        }[row_index]

    for row_index in range(DATA_START_ROW, signature_row - 4):
        worksheet[f"B{row_index}"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        worksheet[f"D{row_index}"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        worksheet.row_dimensions[row_index].height = 18

    _shrink_day_columns(worksheet, signature_row)

    for row_index in range(signature_row - 4, signature_row - 1):
        for cell in worksheet[row_index]:
            cell.font = Font(name="Arial", size=8, bold=True)
        worksheet[f"B{row_index}"].alignment = Alignment(horizontal="left", vertical="center")

    for row_index in range(DATA_START_ROW, signature_row - 1):
        worksheet[f"{AMOUNT_TOTAL_COLUMN}{row_index}"].number_format = "0.00"

    signature_cell = worksheet[f"A{signature_row}"]
    signature_cell.font = Font(name="Arial", size=10)
    signature_cell.alignment = Alignment(horizontal="left", vertical="center")
    worksheet.row_dimensions[signature_row].height = 24


def _shrink_day_columns(worksheet, signature_row: int) -> None:
    for column_index in range(DAY_COLUMN_START, DAY_COLUMN_START + MAX_DAY_COUNT):
        column_letter = get_column_letter(column_index)
        if worksheet.column_dimensions[column_letter].hidden:
            continue
        for row_index in range(HEADER_ROW, signature_row - 1):
            worksheet[f"{column_letter}{row_index}"].alignment = Alignment(
                horizontal="center",
                vertical="center",
                shrink_to_fit=True,
            )


def _safe_sheet_title(value: str) -> str:
    cleaned = "".join("_" if char in "\\/*?:[]" else char for char in value)
    return cleaned[:31]
