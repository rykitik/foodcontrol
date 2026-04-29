from __future__ import annotations

from copy import copy
from functools import lru_cache
from io import BytesIO
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import range_boundaries
from openpyxl.worksheet.worksheet import Worksheet

from .template_fillers import (
    populate_cost_calculation_worksheet,
    populate_cost_statement_worksheet,
    populate_meal_sheet_worksheet,
)
from .template_config import (
    CostCalculationTemplateConfig,
    CostStatementTemplateConfig,
    MealSheetTemplateConfig,
    resolve_cost_calculation_template,
    resolve_cost_statement_template,
    resolve_meal_sheet_template,
)

TEMPLATE_WORKBOOK_NAME = "Питание_2025.xlsm"


def build_cost_statement_template_workbook(payload: dict) -> tuple[Workbook, CostStatementTemplateConfig]:
    config = resolve_cost_statement_template(payload["category_code"])
    worksheet = _load_template_sheet(config.sheet_name, config.visible_range)
    _prepare_template_worksheet(worksheet, config)
    populate_cost_statement_worksheet(worksheet, payload, config)
    workbook = _crop_to_visible_workbook(worksheet, config.visible_range)
    workbook.active.title = _safe_sheet_title(_worksheet_title(payload, "cost_statement"))
    return workbook, config


def build_cost_calculation_template_workbook(
    payload: dict,
    custom_values: dict[str, str] | None = None,
) -> tuple[Workbook, CostCalculationTemplateConfig]:
    config = resolve_cost_calculation_template(payload["category_code"])
    worksheet = _load_template_sheet(config.sheet_name, config.visible_range)
    _prepare_template_worksheet(worksheet, config)
    populate_cost_calculation_worksheet(worksheet, payload, config, custom_values=custom_values)
    workbook = _crop_to_visible_workbook(worksheet, config.visible_range)
    workbook.active.title = _safe_sheet_title(_worksheet_title(payload, "cost_calculation"))
    return workbook, config


def build_meal_sheet_template_workbook(payload: dict) -> tuple[Workbook, MealSheetTemplateConfig]:
    config = resolve_meal_sheet_template(payload["category_code"], payload["meal_type"])
    worksheet = _load_template_sheet(config.sheet_name, config.visible_range)
    _prepare_template_worksheet(worksheet, config)
    populate_meal_sheet_worksheet(worksheet, payload, config)
    workbook = _crop_to_visible_workbook(worksheet, config.visible_range)
    workbook.active.title = _safe_sheet_title(_worksheet_title(payload, "meal_sheet"))
    return workbook, config


def save_workbook_bytes(workbook: Workbook) -> bytes:
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def _load_template_sheet(sheet_name: str, visible_range: str) -> Worksheet:
    workbook = load_workbook(BytesIO(_template_sheet_workbook_bytes(sheet_name, visible_range)), data_only=False)
    return workbook[sheet_name]


def _resolve_template_workbook_path() -> Path:
    candidate = Path(__file__).resolve().parents[2] / "resources" / "accounting_documents" / TEMPLATE_WORKBOOK_NAME
    if candidate.exists():
        return candidate

    raise FileNotFoundError(f"Не найден файл шаблона бухгалтерских форм: {candidate}")


@lru_cache(maxsize=1)
def _template_workbook_bytes() -> bytes:
    return _resolve_template_workbook_path().read_bytes()


@lru_cache(maxsize=None)
def _template_sheet_workbook_bytes(sheet_name: str, visible_range: str) -> bytes:
    source_workbook = load_workbook(BytesIO(_template_workbook_bytes()), data_only=False)
    source_worksheet = source_workbook[sheet_name]
    return save_workbook_bytes(_crop_to_visible_workbook(source_worksheet, visible_range))


def _prepare_template_worksheet(worksheet: Worksheet, config) -> None:
    worksheet.sheet_view.showGridLines = False
    worksheet.print_area = config.visible_range
    _clear_formula_values(worksheet, config.visible_range)

    for clear_range in config.clear_ranges:
        _clear_values(worksheet, clear_range)
        start_row, end_row = _row_bounds(clear_range)
        _unhide_rows(worksheet, start_row, end_row)

    for cell_ref in config.clear_cells:
        worksheet[cell_ref].value = None

    for start_row, end_row in config.unhide_row_spans:
        _unhide_rows(worksheet, start_row, end_row)

    for start_row, end_row in getattr(config, "hide_row_spans", ()):
        _hide_rows(worksheet, start_row, end_row)


def _crop_to_visible_workbook(source: Worksheet, visible_range: str) -> Workbook:
    min_col, min_row, max_col, max_row = range_boundaries(visible_range)
    workbook = Workbook()
    target = workbook.active
    target.title = _safe_sheet_title(source.title)

    _copy_sheet_settings(source, target)

    for source_row in range(min_row, max_row + 1):
        _copy_row_dimension(source, target, source_row, source_row)

    for source_col in range(min_col, max_col + 1):
        _copy_column_dimension(source, target, source_col, source_col)

    for source_row in range(min_row, max_row + 1):
        for source_col in range(min_col, max_col + 1):
            source_cell = source.cell(source_row, source_col)
            if isinstance(source_cell, MergedCell):
                continue
            target_cell = target.cell(source_row, source_col)
            _copy_cell(source_cell, target_cell)

    for merged_range in source.merged_cells.ranges:
        merged_min_col, merged_min_row, merged_max_col, merged_max_row = range_boundaries(str(merged_range))
        if merged_min_col < min_col or merged_max_col > max_col or merged_min_row < min_row or merged_max_row > max_row:
            continue
        target.merge_cells(
            start_row=merged_min_row,
            start_column=merged_min_col,
            end_row=merged_max_row,
            end_column=merged_max_col,
        )

    target.sheet_view.showGridLines = False
    target.print_area = visible_range
    return workbook


def _copy_sheet_settings(source: Worksheet, target: Worksheet) -> None:
    target.sheet_view.showGridLines = source.sheet_view.showGridLines
    target.sheet_view.view = source.sheet_view.view
    target.sheet_view.zoomScale = source.sheet_view.zoomScale
    target.sheet_view.zoomScaleNormal = source.sheet_view.zoomScaleNormal
    target.sheet_view.zoomScalePageLayoutView = source.sheet_view.zoomScalePageLayoutView
    target.page_setup.orientation = source.page_setup.orientation
    target.page_setup.paperSize = source.page_setup.paperSize
    target.page_setup.fitToWidth = source.page_setup.fitToWidth
    target.page_setup.fitToHeight = source.page_setup.fitToHeight
    target.page_setup.scale = source.page_setup.scale
    target.page_setup.firstPageNumber = source.page_setup.firstPageNumber
    target.page_setup.useFirstPageNumber = source.page_setup.useFirstPageNumber
    target.page_setup.pageOrder = source.page_setup.pageOrder
    target.page_setup.blackAndWhite = source.page_setup.blackAndWhite
    target.page_setup.draft = source.page_setup.draft
    target.page_setup.cellComments = source.page_setup.cellComments
    target.page_setup.errors = source.page_setup.errors
    target.page_setup.horizontalDpi = source.page_setup.horizontalDpi
    target.page_setup.verticalDpi = source.page_setup.verticalDpi
    target.page_setup.copies = source.page_setup.copies
    target.page_margins = copy(source.page_margins)
    target.sheet_properties = copy(source.sheet_properties)
    target.print_options = copy(source.print_options)
    target.sheet_format = copy(source.sheet_format)
    if source.print_title_rows:
        target.print_title_rows = source.print_title_rows
    if source.print_title_cols:
        target.print_title_cols = source.print_title_cols


def _copy_row_dimension(source: Worksheet, target: Worksheet, source_row: int, target_row: int) -> None:
    source_dimension = source.row_dimensions[source_row]
    target_dimension = target.row_dimensions[target_row]
    target_dimension.height = source_dimension.height
    target_dimension.hidden = bool(source_dimension.hidden)


def _copy_column_dimension(source: Worksheet, target: Worksheet, source_col: int, target_col: int) -> None:
    source_letter = get_column_letter(source_col)
    target_letter = get_column_letter(target_col)
    source_dimension = source.column_dimensions[source_letter]
    target_dimension = target.column_dimensions[target_letter]
    target_dimension.width = source_dimension.width
    target_dimension.hidden = bool(source_dimension.hidden)


def _copy_cell(source_cell, target_cell) -> None:
    target_cell.value = source_cell.value
    if source_cell.has_style:
        target_cell._style = copy(source_cell._style)
    if source_cell.number_format:
        target_cell.number_format = source_cell.number_format
    if source_cell.font:
        target_cell.font = copy(source_cell.font)
    if source_cell.fill:
        target_cell.fill = copy(source_cell.fill)
    if source_cell.border:
        target_cell.border = copy(source_cell.border)
    if source_cell.alignment:
        target_cell.alignment = copy(source_cell.alignment)
    if source_cell.protection:
        target_cell.protection = copy(source_cell.protection)


def _iter_range_cells(worksheet: Worksheet, range_string: str):
    min_col, min_row, max_col, max_row = range_boundaries(range_string)
    yield from worksheet.iter_rows(
        min_row=min_row,
        max_row=max_row,
        min_col=min_col,
        max_col=max_col,
    )


def _row_bounds(range_string: str) -> tuple[int, int]:
    _, min_row, _, max_row = range_boundaries(range_string)
    return min_row, max_row


def _unhide_rows(worksheet: Worksheet, start_row: int, end_row: int) -> None:
    for row_index in range(start_row, end_row + 1):
        worksheet.row_dimensions[row_index].hidden = False


def _hide_rows(worksheet: Worksheet, start_row: int, end_row: int) -> None:
    for row_index in range(start_row, end_row + 1):
        worksheet.row_dimensions[row_index].hidden = True


def _clear_values(worksheet: Worksheet, range_string: str) -> None:
    for row in _iter_range_cells(worksheet, range_string):
        for cell in row:
            if isinstance(cell, MergedCell):
                continue
            cell.value = None


def _clear_formula_values(worksheet: Worksheet, range_string: str) -> None:
    for row in _iter_range_cells(worksheet, range_string):
        for cell in row:
            if isinstance(cell, MergedCell):
                continue
            if isinstance(cell.value, str) and cell.value.startswith("="):
                cell.value = None


def _safe_sheet_title(value: str) -> str:
    cleaned = "".join("_" if char in "\\/*?:[]" else char for char in value)
    return cleaned[:31]


def _worksheet_title(payload: dict, document_type: str) -> str:
    category_name = str(payload.get("category_name") or "документ")
    if document_type == "meal_sheet":
        meal_label = str(payload.get("meal_type_label") or payload.get("meal_type") or "").strip()
        return f"табель {category_name} {meal_label}".strip()
    if document_type == "cost_statement":
        return f"ведомость {category_name}"
    return f"расчет {category_name}"
