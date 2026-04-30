from __future__ import annotations

from html import escape

from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import range_boundaries

from .cell_values import render_cell_value_html
from .editable_metadata import build_document_editable_metadata
from .template_workbook import (
    build_cost_calculation_template_workbook,
    build_cost_statement_template_workbook,
    build_meal_sheet_template_workbook,
)
from .worksheet_layout import (
    merged_maps,
    overflow_visible_cells,
    printable_height_mm,
    row_height_mm,
    worksheet_uses_fit_to_height,
    worksheet_widths,
)
from .worksheet_styles import render_worksheet_cell_content_style, render_worksheet_cell_style


def render_accounting_document(
    payload: dict,
    workbook,
    config,
    *,
    custom_values: dict[str, str] | None = None,
) -> dict:
    worksheet = workbook.active
    return {
        "title": payload["title"],
        "subtitle": payload["subtitle"],
        "html": _render_worksheet_page(worksheet, config.visible_range),
        "editable_metadata": build_document_editable_metadata(worksheet, payload, config, custom_values),
        "print_mode": "embedded",
        "page_orientation": config.page_orientation,
    }


def render_meal_sheet_document(payload: dict) -> dict:
    workbook, config = build_meal_sheet_template_workbook(payload)
    return render_accounting_document(payload, workbook, config)


def render_cost_statement_document(payload: dict) -> dict:
    workbook, config = build_cost_statement_template_workbook(payload)
    return render_accounting_document(payload, workbook, config)


def render_cost_calculation_document(payload: dict) -> dict:
    workbook, config = build_cost_calculation_template_workbook(payload)
    return render_accounting_document(payload, workbook, config)


def _render_worksheet_page(sheet, visible_range: str) -> str:
    min_col, min_row, max_col, max_row = range_boundaries(visible_range)
    visible_columns = [
        index for index in range(min_col, max_col + 1) if not sheet.column_dimensions[get_column_letter(index)].hidden
    ]
    visible_rows = [index for index in range(min_row, max_row + 1) if not sheet.row_dimensions[index].hidden]
    merged_tops, merged_children = merged_maps(sheet, visible_rows, visible_columns)
    overflow_cells = overflow_visible_cells(sheet, visible_rows, visible_columns, merged_tops, merged_children)

    screen_column_widths, print_column_widths, screen_width, print_width, printable_width = worksheet_widths(
        sheet,
        visible_columns,
    )
    row_heights = {row_index: row_height_mm(sheet, row_index) for row_index in visible_rows}
    screen_height = sum(row_heights.values())
    print_scale = _effective_print_scale(
        sheet,
        screen_width=screen_width,
        screen_height=screen_height,
        print_width=print_width,
    )
    if screen_width > 0 and print_scale < 1:
        print_width = screen_width * print_scale
        print_column_widths = {index: width * print_scale for index, width in screen_column_widths.items()}

    colgroup_html = "".join(
        (
            f'<col style="--accounting-screen-col-width:{screen_column_widths[index]:.4f}mm;'
            f'--accounting-print-col-width:{print_column_widths[index]:.4f}mm;'
            'width:var(--accounting-screen-col-width)" />'
        )
        for index in visible_columns
    )

    rows_html: list[str] = []
    for row_index in visible_rows:
        row_height = row_heights[row_index]
        print_row_height = row_height * print_scale
        cell_html_parts: list[str] = []

        for column_index in visible_columns:
            coordinate = (row_index, column_index)
            if coordinate in merged_children:
                continue

            attrs: list[str] = []

            if coordinate in merged_tops:
                rowspan, colspan = merged_tops[coordinate]
                if rowspan > 1:
                    attrs.append(f'rowspan="{rowspan}"')
                if colspan > 1:
                    attrs.append(f'colspan="{colspan}"')

            cell = sheet.cell(row=row_index, column=column_index)
            attrs.append(f'data-accounting-cell="{escape(cell.coordinate, quote=True)}"')
            overflow_direction = overflow_cells.get(coordinate)
            if overflow_direction is not None:
                attrs.append(f'data-accounting-overflow="{escape(overflow_direction, quote=True)}"')
            cell_style = render_worksheet_cell_style(
                cell,
                allow_overflow=overflow_direction is not None,
                print_scale=print_scale,
            )
            cell_content_html = render_cell_value_html(cell.value, cell.number_format)
            content_style = render_worksheet_cell_content_style(cell, overflow_direction=overflow_direction)
            if content_style:
                cell_content_html = (
                    f'<span style="{escape(content_style, quote=True)}">{cell_content_html}</span>'
                )
            attr_string = f' style="{escape(cell_style, quote=True)}"' if cell_style else ""
            tag_attrs = (" " + " ".join(attrs)) if attrs else ""
            cell_html_parts.append(
                f"<td{tag_attrs}{attr_string}>{cell_content_html}</td>"
            )

        rows_html.append(
            f'<tr data-accounting-row="{row_index}" data-accounting-row-height-mm="{row_height:.2f}" '
            f'style="--accounting-screen-row-height:{row_height:.4f}mm;'
            f'--accounting-print-row-height:{print_row_height:.4f}mm;'
            f'height:var(--accounting-screen-row-height)">{"".join(cell_html_parts)}</tr>'
        )

    print_height = screen_height * print_scale
    worksheet_vars = (
        f"--accounting-screen-width:{screen_width:.4f}mm;"
        f"--accounting-print-width:{print_width:.4f}mm;"
        f"--accounting-screen-height:{screen_height:.4f}mm;"
        f"--accounting-print-height:{print_height:.4f}mm;"
        f"--accounting-print-scale:{print_scale:.8f};"
    )

    return (
        f'<section class="accounting-worksheet-page" style="{worksheet_vars}">'
        + (
            f'<div class="accounting-worksheet" '
            'style="width:var(--accounting-screen-width)" '
            f'data-screen-width-mm="{screen_width:.4f}" '
            f'data-print-width-mm="{print_width:.4f}" '
            f'data-printable-width-mm="{printable_width:.4f}" '
            f'data-print-scale="{print_scale:.8f}">'
        )
        + '<table class="accounting-worksheet-table">'
        + f"<colgroup>{colgroup_html}</colgroup>"
        + f"<tbody>{''.join(rows_html)}</tbody>"
        + "</table>"
        + "</div>"
        + "</section>"
    )


def _effective_print_scale(sheet, *, screen_width: float, screen_height: float, print_width: float) -> float:
    if screen_width <= 0:
        return 1.0

    width_scale = min(print_width / screen_width, 1.0)
    if not worksheet_uses_fit_to_height(sheet) or screen_height <= 0:
        return width_scale

    height_scale = min(printable_height_mm(sheet) / screen_height, 1.0)
    return min(width_scale, height_scale, 1.0)
