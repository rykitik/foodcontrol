from __future__ import annotations

from datetime import datetime
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from app.services.cashier_summary import build_cashier_daily_summary

MONTH_NAMES = {
    1: "ЯНВАРЬ",
    2: "ФЕВРАЛЬ",
    3: "МАРТ",
    4: "АПРЕЛЬ",
    5: "МАЙ",
    6: "ИЮНЬ",
    7: "ИЮЛЬ",
    8: "АВГУСТ",
    9: "СЕНТЯБРЬ",
    10: "ОКТЯБРЬ",
    11: "НОЯБРЬ",
    12: "ДЕКАБРЬ",
}

THIN_SIDE = Side(style="thin", color="D6DCE5")
TABLE_BORDER = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)
TITLE_FILL = PatternFill(fill_type="solid", fgColor="0F172A")
HEADER_FILL = PatternFill(fill_type="solid", fgColor="E2E8F0")
ACCENT_FILL = PatternFill(fill_type="solid", fgColor="DBEAFE")
MUTED_FILL = PatternFill(fill_type="solid", fgColor="F8FAFC")
TITLE_FONT = Font(color="FFFFFF", bold=True, size=14)
HEADER_FONT = Font(color="0F172A", bold=True)
BODY_FONT = Font(color="0F172A")
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")
RIGHT = Alignment(horizontal="right", vertical="center")
WRAP_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
WRAP_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)


def build_cashier_summary_workbook_bytes(
    *,
    days: int,
    history_building_id: int | None,
    month: int | None = None,
    year: int | None = None,
) -> tuple[bytes, str]:
    payload = build_cashier_daily_summary(
        days=days,
        history_building_id=history_building_id,
        month=month,
        year=year,
    )
    workbook = build_cashier_summary_workbook(payload)
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer.getvalue(), _build_filename(payload)


def build_cashier_summary_workbook(payload: dict) -> Workbook:
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Свод"
    _populate_summary_sheet(summary_sheet, payload)
    return workbook


def _build_filename(payload: dict) -> str:
    if payload["filter"]["mode"] == "month":
        return f"cashier_summary_{payload['filter']['year']}_{int(payload['filter']['month']):02d}.xlsx"
    return f"cashier_summary_{payload['period_start']}_{payload['period_end']}.xlsx"


def _period_label(payload: dict) -> str:
    period_start = datetime.fromisoformat(payload["period_start"]).date()
    period_end = datetime.fromisoformat(payload["period_end"]).date()
    if (
        period_start.day == 1
        and period_start.month == period_end.month
        and period_start.year == period_end.year
    ):
        return MONTH_NAMES.get(period_end.month, period_end.strftime("%m.%Y"))
    return f"{payload['period_start']} - {payload['period_end']}"


def _apply_cell_style(cell, *, fill=None, font=None, alignment=None, border: Border | None = TABLE_BORDER) -> None:
    if fill is not None:
        cell.fill = fill
    if font is not None:
        cell.font = font
    if alignment is not None:
        cell.alignment = alignment
    if border is not None:
        cell.border = border


def _format_amount_cell(cell, value: float) -> None:
    cell.value = float(value)
    cell.number_format = '#,##0.00'


def _populate_summary_sheet(sheet, payload: dict) -> None:
    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = "C6"
    sheet.column_dimensions["A"].width = 3
    sheet.column_dimensions["B"].width = 3
    sheet.column_dimensions["C"].width = 34
    sheet.column_dimensions["D"].width = 18
    sheet.column_dimensions["E"].width = 12
    sheet.column_dimensions["F"].width = 12
    sheet.column_dimensions["G"].width = 14
    sheet.column_dimensions["H"].width = 16
    sheet.column_dimensions["I"].width = 4
    sheet.column_dimensions["J"].width = 20
    sheet.column_dimensions["K"].width = 12
    sheet.column_dimensions["L"].width = 12
    sheet.column_dimensions["M"].width = 14

    sheet.merge_cells("C1:H1")
    title_cell = sheet["C1"]
    title_cell.value = "СВОД ПО ПИТАНИЮ"
    _apply_cell_style(title_cell, fill=TITLE_FILL, font=TITLE_FONT, alignment=CENTER)

    sheet["C2"] = "Период"
    sheet["D2"] = _period_label(payload)
    sheet["F2"] = "Сформировано"
    sheet["G2"] = datetime.now().strftime("%Y-%m-%d %H:%M")
    for ref in ("C2", "D2", "F2", "G2"):
        _apply_cell_style(sheet[ref], fill=MUTED_FILL, font=HEADER_FONT if ref in {"C2", "F2"} else BODY_FONT, alignment=LEFT)

    sheet["C3"] = "Все корпуса"
    _apply_cell_style(sheet["C3"], font=HEADER_FONT, alignment=LEFT, border=None)

    for ref, value in {
        "C5": "Корпус",
        "D5": "Тип питания",
        "E5": "цена",
        "F5": "кол-во",
        "G5": "сумма",
        "H5": "ИТОГО",
        "J5": "Тип питания",
        "K5": "цена",
        "L5": "кол-во",
        "M5": "сумма",
    }.items():
        sheet[ref] = value
        _apply_cell_style(sheet[ref], fill=HEADER_FILL, font=HEADER_FONT, alignment=CENTER)

    sheet.merge_cells("J3:M3")
    sheet["J3"] = "Общий срез по всем корпусам"
    _apply_cell_style(sheet["J3"], fill=ACCENT_FILL, font=HEADER_FONT, alignment=CENTER)

    row_index = 6
    for building in payload["buildings_table"]["rows"]:
        line_items = building["line_items"] or [
            {
                "meal_type_label": "Нет выдач",
                "price": None,
                "count": 0,
                "amount": 0.0,
            }
        ]
        start_row = row_index

        for line_item in line_items:
            sheet.cell(row=row_index, column=4, value=line_item["meal_type_label"])
            _apply_cell_style(sheet.cell(row=row_index, column=4), font=BODY_FONT, alignment=LEFT)

            price_cell = sheet.cell(row=row_index, column=5)
            price_cell.value = "" if line_item["price"] is None else float(line_item["price"])
            if line_item["price"] is not None:
                price_cell.number_format = '#,##0.00'
            _apply_cell_style(price_cell, font=BODY_FONT, alignment=RIGHT)

            count_cell = sheet.cell(row=row_index, column=6, value=int(line_item["count"]))
            _apply_cell_style(count_cell, font=BODY_FONT, alignment=RIGHT)

            amount_cell = sheet.cell(row=row_index, column=7)
            _format_amount_cell(amount_cell, float(line_item["amount"]))
            _apply_cell_style(amount_cell, font=BODY_FONT, alignment=RIGHT)

            row_index += 1

        end_row = row_index - 1
        sheet.merge_cells(start_row=start_row, start_column=3, end_row=end_row, end_column=3)
        building_cell = sheet.cell(row=start_row, column=3, value=building["building_name"])
        _apply_cell_style(building_cell, fill=MUTED_FILL, font=HEADER_FONT, alignment=WRAP_LEFT)

        sheet.merge_cells(start_row=start_row, start_column=8, end_row=end_row, end_column=8)
        total_cell = sheet.cell(row=start_row, column=8, value=float(building["total_amount"]))
        total_cell.number_format = '#,##0.00'
        _apply_cell_style(total_cell, fill=MUTED_FILL, font=HEADER_FONT, alignment=WRAP_CENTER)

    totals_row = row_index
    sheet.cell(row=totals_row, column=5, value="ИТОГО")
    _apply_cell_style(sheet.cell(row=totals_row, column=5), fill=ACCENT_FILL, font=HEADER_FONT, alignment=CENTER)
    sheet.cell(row=totals_row, column=6, value=int(payload["buildings_table"]["totals"]["total_count"]))
    _apply_cell_style(sheet.cell(row=totals_row, column=6), fill=ACCENT_FILL, font=HEADER_FONT, alignment=RIGHT)
    _format_amount_cell(sheet.cell(row=totals_row, column=7), float(payload["buildings_table"]["totals"]["total_amount"]))
    _apply_cell_style(sheet.cell(row=totals_row, column=7), fill=ACCENT_FILL, font=HEADER_FONT, alignment=RIGHT)

    total_breakdown_row = 6
    for line_item in payload["buildings_table"]["totals"]["line_items"]:
        sheet.cell(row=total_breakdown_row, column=10, value=line_item["meal_type_label"])
        _apply_cell_style(sheet.cell(row=total_breakdown_row, column=10), font=BODY_FONT, alignment=LEFT)

        total_price_cell = sheet.cell(row=total_breakdown_row, column=11, value=float(line_item["price"]))
        total_price_cell.number_format = '#,##0.00'
        _apply_cell_style(total_price_cell, font=BODY_FONT, alignment=RIGHT)

        total_count_cell = sheet.cell(row=total_breakdown_row, column=12, value=int(line_item["count"]))
        _apply_cell_style(total_count_cell, font=BODY_FONT, alignment=RIGHT)

        total_amount_cell = sheet.cell(row=total_breakdown_row, column=13)
        _format_amount_cell(total_amount_cell, float(line_item["amount"]))
        _apply_cell_style(total_amount_cell, font=BODY_FONT, alignment=RIGHT)
        total_breakdown_row += 1

    sheet.cell(row=total_breakdown_row, column=11, value="ИТОГО")
    _apply_cell_style(sheet.cell(row=total_breakdown_row, column=11), fill=ACCENT_FILL, font=HEADER_FONT, alignment=CENTER)
    sheet.cell(row=total_breakdown_row, column=12, value=int(payload["buildings_table"]["totals"]["total_count"]))
    _apply_cell_style(sheet.cell(row=total_breakdown_row, column=12), fill=ACCENT_FILL, font=HEADER_FONT, alignment=RIGHT)
    total_amount_cell = sheet.cell(row=total_breakdown_row, column=13)
    _format_amount_cell(total_amount_cell, float(payload["buildings_table"]["totals"]["total_amount"]))
    _apply_cell_style(total_amount_cell, fill=ACCENT_FILL, font=HEADER_FONT, alignment=RIGHT)
