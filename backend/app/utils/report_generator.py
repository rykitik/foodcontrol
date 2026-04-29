from __future__ import annotations

from datetime import date
from decimal import Decimal, ROUND_HALF_UP
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

FORM_METADATA = {
    "institution": "МЦК - ЧЭМК Минобразования Чувашии",
    "okud": "0504403",
    "okpo": "02529880",
    "okei": "0383",
    "funding": "Субсидии на иные цели (регион)",
    "department": "Столовая / бухгалтерия",
    "director_name": "Н.Ю. Каргин",
    "chief_accountant_name": "М.М. Суткова",
    "accountant_name": "А.И. Ржанова",
}

MONTHS_RU = {
    1: "январь",
    2: "февраль",
    3: "март",
    4: "апрель",
    5: "май",
    6: "июнь",
    7: "июль",
    8: "август",
    9: "сентябрь",
    10: "октябрь",
    11: "ноябрь",
    12: "декабрь",
}


def format_russian_date(value):
    return value.strftime("%d.%m.%Y")


def russian_month_name(value: date) -> str:
    return MONTHS_RU[value.month]


def rubles_to_words(amount: float) -> str:
    units_male = (
        "",
        "один",
        "два",
        "три",
        "четыре",
        "пять",
        "шесть",
        "семь",
        "восемь",
        "девять",
    )
    units_female = (
        "",
        "одна",
        "две",
        "три",
        "четыре",
        "пять",
        "шесть",
        "семь",
        "восемь",
        "девять",
    )
    teens = (
        "десять",
        "одиннадцать",
        "двенадцать",
        "тринадцать",
        "четырнадцать",
        "пятнадцать",
        "шестнадцать",
        "семнадцать",
        "восемнадцать",
        "девятнадцать",
    )
    tens = (
        "",
        "",
        "двадцать",
        "тридцать",
        "сорок",
        "пятьдесят",
        "шестьдесят",
        "семьдесят",
        "восемьдесят",
        "девяносто",
    )
    hundreds = (
        "",
        "сто",
        "двести",
        "триста",
        "четыреста",
        "пятьсот",
        "шестьсот",
        "семьсот",
        "восемьсот",
        "девятьсот",
    )
    groups = (
        ("рубль", "рубля", "рублей", False),
        ("тысяча", "тысячи", "тысяч", True),
        ("миллион", "миллиона", "миллионов", False),
    )

    def plural_form(number: int, forms) -> str:
        number = abs(number) % 100
        if 10 < number < 20:
            return forms[2]
        number = number % 10
        if number == 1:
            return forms[0]
        if 1 < number < 5:
            return forms[1]
        return forms[2]

    def triplet_to_words(number: int, female: bool) -> list[str]:
        result = []
        result.append(hundreds[number // 100])
        remainder = number % 100
        if 10 <= remainder <= 19:
            result.append(teens[remainder - 10])
        else:
            result.append(tens[remainder // 10])
            units = units_female if female else units_male
            result.append(units[remainder % 10])
        return [item for item in result if item]

    value = Decimal(str(amount)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    rubles = int(value)
    kopeks = int((value - rubles) * 100)
    if rubles == 0:
        words = ["ноль", plural_form(0, groups[0][:3])]
    else:
        words = []
        group_index = 0
        while rubles > 0 and group_index < len(groups):
            rubles, triplet = divmod(rubles, 1000)
            if triplet:
                forms = groups[group_index]
                words = triplet_to_words(triplet, forms[3]) + [plural_form(triplet, forms[:3])] + words
            group_index += 1
    return f"{' '.join(words).strip().capitalize()} {kopeks:02d} коп."


def build_ticket_journal_workbook(rows: list[dict], month: int, year: int) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Журнал талонов"

    title = f"Журнал талонов за {month:02d}.{year}"
    headers = [
        "№",
        "Студент",
        "Категория",
        "Корпус питания",
        "Месяц",
        "Период действия",
        "Статус",
        "Выдач",
        "Требует внимания",
    ]

    sheet.merge_cells("A1:I1")
    sheet["A1"] = title
    sheet["A1"].font = Font(size=14, bold=True)
    sheet["A1"].alignment = Alignment(horizontal="center")

    for index, header in enumerate(headers, start=1):
        cell = sheet.cell(row=3, column=index, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_index, row in enumerate(rows, start=4):
        values = [
            row_index - 3,
            row["student_name"],
            row["category_name"],
            row.get("building_name") or row.get("building_id"),
            f'{row["month"]}/{row["year"]}',
            f'{row["start_date"]} - {row["end_date"]}',
            row["status"],
            row.get("meal_records_count", 0),
            "Да" if row.get("requires_attention") else "Нет",
        ]
        for column_index, value in enumerate(values, start=1):
            cell = sheet.cell(row=row_index, column=column_index, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    widths = {
        1: 6,
        2: 34,
        3: 18,
        4: 10,
        5: 12,
        6: 24,
        7: 16,
        8: 10,
        9: 18,
    }
    for column_index, width in widths.items():
        sheet.column_dimensions[get_column_letter(column_index)].width = width

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output.getvalue()
