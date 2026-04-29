from __future__ import annotations

import re
from io import BytesIO

import pytest
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import get_column_letter

from app.models import Category
from app.services.accounting_documents.template_config import resolve_meal_sheet_template

MM_PER_INCH = 25.4

_PREVIEW_BASE_STYLES = """
<style>
  .accounting-worksheet-page {
    display: block;
    width: 100%;
    overflow-x: auto;
  }

  .accounting-worksheet {
    display: block;
    width: var(--accounting-screen-width, auto);
    min-width: var(--accounting-screen-width, auto);
    color: #111111;
  }

  .accounting-worksheet-table {
    width: 100%;
    min-width: 100%;
    border-collapse: collapse;
    table-layout: fixed;
  }

  .accounting-worksheet-table col {
    width: var(--accounting-screen-col-width) !important;
  }

  .accounting-worksheet-table td {
    padding: 0;
    box-sizing: border-box;
  }
</style>
"""


def _login(client, username: str, password: str = "password123") -> dict[str, str]:
    response = client.post("/api/auth/login", json={"username": username, "password": password})
    assert response.status_code == 200
    token = response.get_json()["data"]["token"]
    return {"Authorization": f"Bearer {token}"}


def _category_id_by_code(app, code: str) -> int:
    with app.app_context():
        category = Category.query.filter_by(code=code).first()
        assert category is not None
        return int(category.id)


def _workbook_from_response(response):
    assert response.status_code == 200
    workbook = load_workbook(BytesIO(response.data))
    return workbook.active


def _worksheet_visible_column_widths_mm(sheet) -> list[float]:
    widths: list[float] = []
    for column_index in range(1, sheet.max_column + 1):
        column_letter = get_column_letter(column_index)
        if sheet.column_dimensions[column_letter].hidden:
            continue
        width = float(sheet.column_dimensions[column_letter].width or 8.43)
        pixels = width * 12 if width < 1 else width * 7 + 5
        widths.append(pixels * MM_PER_INCH / 96)
    return widths


def _preview_screen_col_widths_mm(html: str) -> list[float]:
    return [
        float(width)
        for width in re.findall(r"--accounting-screen-col-width:([0-9.]+)mm;--accounting-print-col-width:", html)
    ]


def _preview_page_html(preview_html: str) -> str:
    return (
        "<!doctype html><html><head><meta charset='UTF-8'>"
        + _PREVIEW_BASE_STYLES
        + "</head><body>"
        + preview_html
        + "</body></html>"
    )


def _browser_col_widths_mm_from_preview(preview_html: str) -> list[float]:
    sync_api = pytest.importorskip("playwright.sync_api", reason="Playwright is not installed")
    with sync_api.sync_playwright() as playwright:
        try:
            browser = playwright.chromium.launch(headless=True)
        except Exception as exc:  # pragma: no cover - environment dependent
            pytest.skip(f"Playwright browser is unavailable: {exc}")

        page = browser.new_page(viewport={"width": 1920, "height": 1200})
        page.set_content(_preview_page_html(preview_html), wait_until="load")
        pixel_widths = page.eval_on_selector_all(
            ".accounting-worksheet-table col",
            "cols => cols.map((col) => Number.parseFloat(getComputedStyle(col).width) || col.getBoundingClientRect().width)",
        )
        browser.close()

    return [float(width) * MM_PER_INCH / 96 for width in pixel_widths]


def _find_day_column(sheet, *, header_row: int, day_number: int) -> str:
    for cell in sheet[header_row]:
        if isinstance(cell, MergedCell):
            continue
        if cell.value == day_number:
            return cell.column_letter
    raise AssertionError(f"Day column {day_number} was not found")


def _visible_columns(sheet) -> list[str]:
    columns: list[str] = []
    for column_index in range(1, sheet.max_column + 1):
        column_letter = get_column_letter(column_index)
        if sheet.column_dimensions[column_letter].hidden:
            continue
        columns.append(column_letter)
    return columns


def _assert_preview_and_browser_widths_match_workbook(
    *,
    preview_html: str,
    sheet,
    preview_tolerance_mm: float = 0.2,
    browser_tolerance_mm: float = 0.35,
) -> list[float]:
    preview_widths_mm = _preview_screen_col_widths_mm(preview_html)
    workbook_widths_mm = _worksheet_visible_column_widths_mm(sheet)
    browser_widths_mm = _browser_col_widths_mm_from_preview(preview_html)

    assert len(preview_widths_mm) == len(workbook_widths_mm) == len(browser_widths_mm)
    for preview_width, workbook_width in zip(preview_widths_mm, workbook_widths_mm, strict=True):
        assert abs(preview_width - workbook_width) <= preview_tolerance_mm
    for browser_width, workbook_width in zip(browser_widths_mm, workbook_widths_mm, strict=True):
        assert abs(browser_width - workbook_width) <= browser_tolerance_mm

    return browser_widths_mm


def test_accountant_meal_sheet_preview_browser_dom_widths_match_workbook(client, app):
    accountant_headers = _login(client, "accountant")
    category_id = _category_id_by_code(app, "ovz")
    config = resolve_meal_sheet_template("ovz", "lunch")
    request_payload = {
        "month": 2,
        "year": 2025,
        "category_id": category_id,
        "meal_type": "lunch",
    }

    document_response = client.post(
        "/api/reports/accounting-documents/meal-sheet/document",
        headers=accountant_headers,
        json=request_payload,
    )
    assert document_response.status_code == 200
    document_payload = document_response.get_json()["data"]

    workbook_response = client.post(
        "/api/reports/accounting-documents/meal-sheet/xlsx",
        headers=accountant_headers,
        json=request_payload,
    )
    sheet = _workbook_from_response(workbook_response)

    browser_widths_mm = _assert_preview_and_browser_widths_match_workbook(
        preview_html=document_payload["html"],
        sheet=sheet,
    )

    visible_columns = _visible_columns(sheet)
    index_column_position = visible_columns.index(config.index_column)
    day_8_column = _find_day_column(sheet, header_row=config.header_day_row, day_number=8)
    day_8_position = visible_columns.index(day_8_column)

    assert browser_widths_mm[index_column_position] < browser_widths_mm[day_8_position]


def test_accountant_combined_meal_sheet_preview_browser_dom_widths_match_workbook(client, app):
    accountant_headers = _login(client, "accountant")
    category_id = _category_id_by_code(app, "ovz")
    request_payload = {
        "month": 2,
        "year": 2025,
        "category_id": category_id,
    }

    document_response = client.post(
        "/api/reports/accounting-documents/combined-meal-sheet/document",
        headers=accountant_headers,
        json=request_payload,
    )
    assert document_response.status_code == 200
    document_payload = document_response.get_json()["data"]
    assert document_payload["page_orientation"] == "landscape"

    workbook_response = client.post(
        "/api/reports/accounting-documents/combined-meal-sheet/xlsx",
        headers=accountant_headers,
        json=request_payload,
    )
    sheet = _workbook_from_response(workbook_response)

    _assert_preview_and_browser_widths_match_workbook(
        preview_html=document_payload["html"],
        sheet=sheet,
    )


def test_accountant_cost_statement_preview_browser_dom_widths_match_workbook(client, app):
    accountant_headers = _login(client, "accountant")
    category_id = _category_id_by_code(app, "low_income")
    request_payload = {
        "month": 2,
        "year": 2025,
        "category_id": category_id,
    }

    document_response = client.post(
        "/api/reports/accounting-documents/cost-statement/document",
        headers=accountant_headers,
        json=request_payload,
    )
    assert document_response.status_code == 200
    document_payload = document_response.get_json()["data"]

    workbook_response = client.post(
        "/api/reports/accounting-documents/cost-statement/xlsx",
        headers=accountant_headers,
        json=request_payload,
    )
    sheet = _workbook_from_response(workbook_response)

    _assert_preview_and_browser_widths_match_workbook(
        preview_html=document_payload["html"],
        sheet=sheet,
    )


def test_accountant_cost_calculation_preview_browser_dom_widths_match_workbook(client, app):
    accountant_headers = _login(client, "accountant")
    category_id = _category_id_by_code(app, "orphan")
    request_payload = {
        "month": 2,
        "year": 2025,
        "category_id": category_id,
    }

    document_response = client.post(
        "/api/reports/accounting-documents/cost-calculation/document",
        headers=accountant_headers,
        json=request_payload,
    )
    assert document_response.status_code == 200
    document_payload = document_response.get_json()["data"]

    workbook_response = client.post(
        "/api/reports/accounting-documents/cost-calculation/xlsx",
        headers=accountant_headers,
        json=request_payload,
    )
    sheet = _workbook_from_response(workbook_response)

    _assert_preview_and_browser_widths_match_workbook(
        preview_html=document_payload["html"],
        sheet=sheet,
    )
