from __future__ import annotations

from calendar import monthrange
from datetime import date, timedelta
from io import BytesIO

from openpyxl import load_workbook

from app.services.cashier_summary_workbook import MONTH_NAMES
from app.models import Student, Ticket, User, db


def login(client, username: str, password: str = "password123") -> dict[str, str]:
    response = client.post("/api/auth/login", json={"username": username, "password": password})
    assert response.status_code == 200
    payload = response.get_json()
    token = payload["data"]["token"]
    return {"Authorization": f"Bearer {token}"}


def create_user(app, *, username: str, role: str, building_id: int | None, password: str = "password123") -> None:
    with app.app_context():
        user = User(
            username=username,
            full_name=f"{username} test",
            role=role,
            building_id=building_id,
            is_active=True,
        )
        user.set_password(password)
        db.session.add(user)
        db.session.commit()


def first_service_day(reference: date | None = None) -> date:
    current = reference or date.today().replace(day=1)
    while current.weekday() == 6:
        current += timedelta(days=1)
    return current


def first_service_day_for_month(year: int, month: int) -> date:
    current = date(year, month, 1)
    while current.weekday() == 6:
        current += timedelta(days=1)
    return current


def previous_month(reference: date) -> tuple[int, int]:
    month = reference.month - 1
    year = reference.year
    if month == 0:
        month = 12
        year -= 1
    return year, month


def create_direct_ticket(app, *, student_id: str, creator_username: str, month: int, year: int) -> str:
    with app.app_context():
        student = Student.query.filter_by(id=student_id).first()
        creator = User.query.filter_by(username=creator_username).first()
        assert student is not None
        assert creator is not None

        ticket = Ticket(
            student_id=student.id,
            category_id=student.category_id,
            month=month,
            year=year,
            start_date=date(year, month, 1),
            end_date=date(year, month, monthrange(year, month)[1]),
            status="active",
            qr_code=f"summary-xlsx-{student.id}-{year}-{month:02d}",
            created_by=creator.id,
        )
        db.session.add(ticket)
        db.session.commit()
        return ticket.id


def create_summary_student(app, *, student_card: str, building_id: int = 1) -> dict:
    with app.app_context():
        template = Student.query.filter_by(building_id=building_id, is_active=True).first()
        assert template is not None

        student = Student(
            full_name=f"Summary XLSX {student_card}",
            student_card=student_card,
            group_name="TEST-001",
            building_id=building_id,
            meal_building_id=building_id,
            allow_all_meal_buildings=False,
            category_id=template.category_id,
            is_active=True,
        )
        db.session.add(student)
        db.session.commit()
        return {
            "id": student.id,
            "meal_type": student.category.meal_types[-1],
        }


def test_cashier_daily_summary_xlsx_contains_summary_sheet(client, app):
    create_user(app, username="cashier2", role="cashier", building_id=2)

    head_headers = login(client, "headsocial")
    cashier1_headers = login(client, "cashier1")
    cashier2_headers = login(client, "cashier2")
    service_day = first_service_day(date.today().replace(day=1))

    students = client.get("/api/students/search", headers=head_headers).get_json()["data"]
    building1_student = next(student for student in students if student["effective_meal_building_id"] == 1)
    building2_student = next(student for student in students if student["effective_meal_building_id"] == 2)

    for student in (building1_student, building2_student):
        ticket_response = client.post(
            "/api/tickets",
            headers=head_headers,
            json={"student_id": student["id"], "month": service_day.month, "year": service_day.year},
        )
        assert ticket_response.status_code in {201, 409}

    assert (
        client.post(
            "/api/meals/record",
            headers=cashier1_headers,
            json={
                "code": building1_student["student_card"],
                "meal_type": building1_student["category"]["meal_types"][-1],
                "issue_date": service_day.isoformat(),
            },
        ).status_code
        == 201
    )
    assert (
        client.post(
            "/api/meals/record",
            headers=cashier2_headers,
            json={
                "code": building2_student["student_card"],
                "meal_type": building2_student["category"]["meal_types"][-1],
                "issue_date": service_day.isoformat(),
            },
        ).status_code
        == 201
    )

    response = client.get("/api/meals/daily-summary/xlsx?days=31", headers=cashier1_headers)

    assert response.status_code == 200
    assert response.mimetype == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    workbook = load_workbook(BytesIO(response.data))
    assert workbook.sheetnames == ["Свод"]

    summary_sheet = workbook["Свод"]

    assert summary_sheet["C1"].value == "СВОД ПО ПИТАНИЮ"
    assert summary_sheet["C3"].value == "Все корпуса"
    assert summary_sheet["C6"].value == "Корпус 1, Ленина, д.9"
    summary_text = " ".join(
        str(cell.value)
        for row in summary_sheet.iter_rows()
        for cell in row
        if cell.value not in {None, ""}
    )
    assert "Корпус 2, Яковлева, д.17" in summary_text
    assert "Общий срез по всем корпусам" in summary_text
    assert service_day.isoformat() not in {str(cell.value) for row in summary_sheet.iter_rows() for cell in row if cell.value not in {None, ""}}


def test_cashier_daily_summary_xlsx_supports_month_filter(client, app):
    cashier_headers = login(client, "cashier1")

    current_reference = date.today().replace(day=1)
    previous_year, previous_month_value = previous_month(current_reference)
    previous_service_day = first_service_day_for_month(previous_year, previous_month_value)
    student = create_summary_student(app, student_card="990002")
    ticket_id = create_direct_ticket(
        app,
        student_id=student["id"],
        creator_username="headsocial",
        month=previous_month_value,
        year=previous_year,
    )

    meal_response = client.post(
        "/api/meals/record",
        headers=cashier_headers,
        json={
            "ticket_id": ticket_id,
            "meal_type": student["meal_type"],
            "issue_date": previous_service_day.isoformat(),
        },
    )
    assert meal_response.status_code == 201

    response = client.get(
        f"/api/meals/daily-summary/xlsx?month={previous_month_value}&year={previous_year}",
        headers=cashier_headers,
    )

    assert response.status_code == 200
    workbook = load_workbook(BytesIO(response.data))
    summary_sheet = workbook["Свод"]

    assert workbook.sheetnames == ["Свод"]
    assert summary_sheet["D2"].value == MONTH_NAMES[previous_month_value]
