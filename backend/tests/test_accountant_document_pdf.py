from __future__ import annotations

from io import BytesIO
import pytest
from openpyxl import Workbook, load_workbook

from app.services.accounting_documents.pdf import (
    XlsxToPdfConversionError,
    _libreoffice_uno_command,
    convert_xlsx_to_pdf,
    inspect_pdf_first_page_geometry,
    prepare_xlsx_for_pdf,
    validate_pdf_page_geometry,
)


def login(client, username: str, password: str = "password123") -> dict[str, str]:
    response = client.post("/api/auth/login", json={"username": username, "password": password})
    assert response.status_code == 200
    token = response.get_json()["data"]["token"]
    return {"Authorization": f"Bearer {token}"}


def meal_sheet_request() -> dict:
    return {
        "month": 2,
        "year": 2025,
        "category_id": 1,
        "meal_type": "breakfast",
    }


def workbook_bytes(workbook: Workbook) -> bytes:
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def landscape_pdf_bytes() -> bytes:
    return b"%PDF-1.4\n1 0 obj << /Type /Page /MediaBox [0 0 842 595] /CropBox [0 0 842 595] >> endobj\n%%EOF"


def portrait_rotated_pdf_bytes() -> bytes:
    return b"%PDF-1.4\n1 0 obj << /Type /Page /MediaBox [0 0 595 842] /Rotate 90 >> endobj\n%%EOF"


def workbook_for_pdf_scale(*, narrow: bool = True) -> Workbook:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Accounting"
    worksheet["A1"] = "Header"
    worksheet["B1"] = "Value"
    worksheet.print_area = "A1:B1" if narrow else "A1:Z1"
    worksheet.page_setup.orientation = "landscape"
    worksheet.page_setup.paperSize = "9"
    worksheet.page_setup.fitToWidth = 1
    worksheet.page_setup.fitToHeight = 0
    worksheet.page_setup.scale = None
    worksheet.sheet_properties.pageSetUpPr.fitToPage = True
    worksheet.page_margins.left = 0.2
    worksheet.page_margins.right = 0.2
    for column_index in range(1, 27):
        column_letter = chr(64 + column_index)
        worksheet.column_dimensions[column_letter].width = 8 if narrow else 30
    return workbook


def test_accounting_pdf_endpoint_requires_accountant_role(client):
    headers = login(client, "social1")

    response = client.post(
        "/api/reports/accounting-documents/meal-sheet/pdf",
        headers=headers,
        json=meal_sheet_request(),
    )

    assert response.status_code == 403


def test_accounting_pdf_endpoint_returns_unavailable_when_disabled(client, app):
    app.config["ACCOUNTING_PDF_ENABLED"] = False
    headers = login(client, "accountant")

    response = client.post(
        "/api/reports/accounting-documents/meal-sheet/pdf",
        headers=headers,
        json=meal_sheet_request(),
    )

    assert response.status_code == 503


def test_accounting_pdf_endpoint_converts_generated_workbook_bytes(client, app, monkeypatch):
    app.config["ACCOUNTING_PDF_ENABLED"] = True
    app.config["ACCOUNTING_PDF_MAX_AUTO_SCALE"] = 115
    headers = login(client, "accountant")
    captured: dict[str, bytes] = {}

    def fake_convert(
        workbook_bytes: bytes,
        *,
        libreoffice_bin: str,
        timeout_seconds: int,
        max_auto_scale_percent: int,
    ) -> bytes:
        captured["workbook_bytes"] = workbook_bytes
        captured["libreoffice_bin"] = libreoffice_bin.encode()
        captured["timeout_seconds"] = str(timeout_seconds).encode()
        captured["max_auto_scale_percent"] = str(max_auto_scale_percent).encode()
        return b"%PDF-1.4\n%%EOF"

    monkeypatch.setattr("app.routes.reports.convert_xlsx_to_pdf", fake_convert)

    response = client.post(
        "/api/reports/accounting-documents/meal-sheet/pdf",
        headers=headers,
        json=meal_sheet_request(),
    )

    assert response.status_code == 200
    assert response.mimetype == "application/pdf"
    assert response.headers["Cache-Control"] == "no-store"
    assert response.headers["Content-Disposition"].startswith("inline;")
    assert response.data.startswith(b"%PDF")
    assert captured["workbook_bytes"].startswith(b"PK")
    assert captured["max_auto_scale_percent"] == b"115"


def test_accounting_pdf_endpoint_handles_conversion_errors(client, app, monkeypatch):
    app.config["ACCOUNTING_PDF_ENABLED"] = True
    headers = login(client, "accountant")

    def fake_convert(*args, **kwargs):
        raise XlsxToPdfConversionError("LibreOffice conversion failed")

    monkeypatch.setattr("app.routes.reports.convert_xlsx_to_pdf", fake_convert)

    response = client.post(
        "/api/reports/accounting-documents/meal-sheet/pdf",
        headers=headers,
        json=meal_sheet_request(),
    )

    assert response.status_code == 503


def test_xlsx_to_pdf_converter_uses_uno_export_and_validates_landscape_pdf(monkeypatch):
    calls: list[dict] = []
    source_workbook = workbook_for_pdf_scale()

    def fake_uno_export(**kwargs):
        calls.append(kwargs)
        kwargs["pdf_path"].write_bytes(landscape_pdf_bytes())

    monkeypatch.setattr("app.services.accounting_documents.pdf._convert_xlsx_to_pdf_with_uno", fake_uno_export)

    pdf_bytes = convert_xlsx_to_pdf(
        workbook_bytes(source_workbook),
        libreoffice_bin="soffice",
        timeout_seconds=12,
        max_auto_scale_percent=125,
    )

    assert pdf_bytes.startswith(b"%PDF")
    assert calls[0]["libreoffice_bin"] == "soffice"
    assert calls[0]["timeout_seconds"] == 12


def test_libreoffice_uno_command_uses_headless_tmp_profile_and_socket_accept(tmp_path):
    command = _libreoffice_uno_command("soffice", profile_dir=tmp_path / "profile", port=23456)

    assert command[:4] == ["soffice", "--headless", "--norestore", "--nofirststartwizard"]
    assert any(part.startswith("-env:UserInstallation=file:///") for part in command)
    assert "--accept=socket,host=127.0.0.1,port=23456;urp;StarOffice.ComponentContext" in command
    assert "--convert-to" not in command


def test_pdf_page_geometry_detects_landscape_without_rotation():
    geometry = inspect_pdf_first_page_geometry(landscape_pdf_bytes())

    assert geometry.is_landscape is True
    assert geometry.rotate == 0
    assert geometry.media_width > geometry.media_height


def test_pdf_page_geometry_rejects_portrait_or_rotated_landscape_expected():
    with pytest.raises(XlsxToPdfConversionError):
        validate_pdf_page_geometry(portrait_rotated_pdf_bytes(), expected_landscape=True)


def test_pdf_copy_auto_scales_narrow_print_area_without_changing_source_workbook():
    source_workbook = workbook_for_pdf_scale(narrow=True)
    source_bytes = workbook_bytes(source_workbook)

    pdf_bytes, diagnostics = prepare_xlsx_for_pdf(source_bytes, max_auto_scale_percent=125)

    original_sheet = load_workbook(BytesIO(source_bytes)).active
    pdf_sheet = load_workbook(BytesIO(pdf_bytes)).active
    diagnostic = diagnostics[0]

    assert diagnostic.applied is True
    assert diagnostic.target_scale_percent == 125
    assert diagnostic.print_area_width_mm < diagnostic.printable_width_mm
    assert original_sheet.page_setup.fitToWidth == 1
    assert original_sheet.page_setup.fitToHeight == 0
    assert original_sheet.page_setup.scale is None
    assert original_sheet.sheet_properties.pageSetUpPr.fitToPage is True
    assert pdf_sheet.page_setup.fitToWidth == 0
    assert pdf_sheet.page_setup.fitToHeight == 0
    assert int(pdf_sheet.page_setup.scale) == 125
    assert pdf_sheet.sheet_properties.pageSetUpPr.fitToPage is False
    assert pdf_sheet["A1"].value == original_sheet["A1"].value
    assert pdf_sheet["B1"].value == original_sheet["B1"].value


def test_pdf_copy_auto_scale_respects_max_percent():
    source_bytes = workbook_bytes(workbook_for_pdf_scale(narrow=True))

    pdf_bytes, diagnostics = prepare_xlsx_for_pdf(source_bytes, max_auto_scale_percent=110)
    pdf_sheet = load_workbook(BytesIO(pdf_bytes)).active

    assert diagnostics[0].applied is True
    assert diagnostics[0].target_scale_percent == 110
    assert int(pdf_sheet.page_setup.scale) == 110


def test_pdf_copy_keeps_fit_to_width_when_print_area_is_wider_than_page():
    source_bytes = workbook_bytes(workbook_for_pdf_scale(narrow=False))

    pdf_bytes, diagnostics = prepare_xlsx_for_pdf(source_bytes, max_auto_scale_percent=125)
    pdf_sheet = load_workbook(BytesIO(pdf_bytes)).active

    assert pdf_bytes == source_bytes
    assert diagnostics[0].applied is False
    assert diagnostics[0].target_scale_percent == 100
    assert diagnostics[0].print_area_width_mm > diagnostics[0].printable_width_mm
    assert pdf_sheet.page_setup.fitToWidth == 1
    assert pdf_sheet.page_setup.fitToHeight == 0
    assert pdf_sheet.page_setup.scale is None
    assert pdf_sheet.sheet_properties.pageSetUpPr.fitToPage is True
