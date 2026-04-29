from __future__ import annotations

from io import BytesIO

from openpyxl import load_workbook

from test_accountant_document_forms import login, prepare_accounting_records


def test_accountant_ovz_cost_calculation_document_and_xlsx(client, app):
    accountant_headers = login(client, "accountant")
    prepared = prepare_accounting_records(client, app, student_card="100002", meal_types=("breakfast", "lunch"))

    response = client.post(
        "/api/reports/accounting-documents/cost-calculation/document",
        headers=accountant_headers,
        json={
            "month": prepared["month"],
            "year": prepared["year"],
            "category_id": prepared["category_id"],
        },
    )

    assert response.status_code == 200
    payload = response.get_json()["data"]
    assert prepared["student"]["full_name"] in payload["html"]
    assert "с ОВЗ колледжа" in payload["html"]

    xlsx_response = client.post(
        "/api/reports/accounting-documents/cost-calculation/xlsx",
        headers=accountant_headers,
        json={
            "month": prepared["month"],
            "year": prepared["year"],
            "category_id": prepared["category_id"],
        },
    )

    assert xlsx_response.status_code == 200
    workbook = load_workbook(BytesIO(xlsx_response.data))
    sheet = workbook.active

    assert sheet["B4"].value == "студентам с ОВЗ колледжа"
    assert sheet["B9"].value == prepared["student"]["full_name"]


def test_accountant_all_categories_cost_calculation_document_and_xlsx(client, app):
    accountant_headers = login(client, "accountant")
    orphan_prepared = prepare_accounting_records(client, app, student_card="100001", meal_types=("lunch",))
    ovz_prepared = prepare_accounting_records(client, app, student_card="100002", meal_types=("breakfast", "lunch"))

    response = client.post(
        "/api/reports/accounting-documents/cost-calculation/document",
        headers=accountant_headers,
        json={
            "month": orphan_prepared["month"],
            "year": orphan_prepared["year"],
            "category_id": 0,
        },
    )

    assert response.status_code == 200
    payload = response.get_json()["data"]
    assert orphan_prepared["student"]["full_name"] in payload["html"]
    assert ovz_prepared["student"]["full_name"] in payload["html"]
    assert "всех категорий" in payload["html"]

    xlsx_response = client.post(
        "/api/reports/accounting-documents/cost-calculation/xlsx",
        headers=accountant_headers,
        json={
            "month": orphan_prepared["month"],
            "year": orphan_prepared["year"],
            "category_id": 0,
        },
    )

    assert xlsx_response.status_code == 200
    workbook = load_workbook(BytesIO(xlsx_response.data))
    sheet = workbook.active

    expected_accrued = round(float(sheet["G7"].value) * float(sheet["C7"].value), 2)

    assert sheet["B4"].value == "студентам всех категорий колледжа"
    assert {sheet["B9"].value, sheet["B10"].value} == {
        orphan_prepared["student"]["full_name"],
        ovz_prepared["student"]["full_name"],
    }
    assert sheet["D9"].value == expected_accrued
    assert sheet["D10"].value == expected_accrued
    assert sheet["E39"].value in {435, 435.0}
