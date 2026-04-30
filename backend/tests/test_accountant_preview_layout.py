from __future__ import annotations

import re

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, PatternFill, Side

from app.models import Category
from app.services.accounting_documents.html import _render_worksheet_page
from app.services.accounting_documents.worksheet_layout import merged_maps, overflow_visible_cells
from app.services.accounting_documents.worksheet_styles import render_worksheet_cell_content_style


def _login(client, username: str, password: str = "password123") -> dict[str, str]:
    response = client.post("/api/auth/login", json={"username": username, "password": password})
    assert response.status_code == 200
    token = response.get_json()["data"]["token"]
    return {"Authorization": f"Bearer {token}"}


def _category_id_by_code(app, code: str) -> int:
    with app.app_context():
        category = Category.query.filter_by(code=code).first()
        assert category is not None
        return int(category.id)


def test_preview_overflow_ignores_cells_with_borders_or_fill():
    workbook = Workbook()
    sheet = workbook.active

    for row_index in range(1, 5):
        sheet.cell(row=row_index, column=1, value=None)
        target = sheet.cell(row=row_index, column=2, value="МЦК - ЧЭМК Минобразования Чувашии")
        target.alignment = Alignment(horizontal="right")

    sheet["B2"].border = Border(left=Side(style="thin"))
    sheet["B3"].fill = PatternFill("solid", fgColor="D9EAF7")
    sheet["A4"].border = Border(bottom=Side(style="thin"))

    visible_rows = [1, 2, 3, 4]
    visible_columns = [1, 2]
    merged_tops, merged_children = merged_maps(sheet, visible_rows, visible_columns)

    overflow = overflow_visible_cells(
        sheet,
        visible_rows,
        visible_columns,
        merged_tops,
        merged_children,
    )

    assert overflow[(1, 2)] == "left"
    assert (2, 2) not in overflow
    assert (3, 2) not in overflow
    assert (4, 2) not in overflow


def test_preview_left_overflow_content_is_anchored_to_source_cell_right_edge():
    workbook = Workbook()
    sheet = workbook.active
    cell = sheet["F7"]
    cell.value = "размер компенсац.выплат в день:"
    cell.alignment = Alignment(horizontal="right")

    style = render_worksheet_cell_content_style(cell, overflow_direction="left")

    assert "display:block" in style
    assert "position:absolute" in style
    assert "right:0" in style
    assert "left:auto" in style
    assert "transform:translateY(-50%)" in style
    assert "text-align:right" in style
    assert "pointer-events:none" in style


def test_preview_center_overflow_content_keeps_center_alignment():
    workbook = Workbook()
    sheet = workbook.active
    cell = sheet["C3"]
    cell.value = "центр"
    cell.alignment = Alignment(horizontal="center")

    style = render_worksheet_cell_content_style(cell, overflow_direction="center")

    assert "left:50%" in style
    assert "transform:translate(-50%, -50%)" in style
    assert "text-align:center" in style


def test_preview_right_overflow_content_is_anchored_to_source_cell_left_edge():
    workbook = Workbook()
    sheet = workbook.active
    cell = sheet["Q7"]
    cell.value = "right overflow"
    cell.alignment = Alignment(horizontal="left")

    style = render_worksheet_cell_content_style(cell, overflow_direction="right")

    assert "position:absolute" in style
    assert "left:0" in style
    assert "right:auto" in style
    assert "transform:translateY(-50%)" in style
    assert "text-align:left" in style
    assert "pointer-events:none" in style


def test_preview_without_overflow_does_not_wrap_cell_content():
    workbook = Workbook()
    sheet = workbook.active
    cell = sheet["AS7"]
    cell.value = "MCK - CHEMK"
    cell.alignment = Alignment(horizontal="right")

    assert render_worksheet_cell_content_style(cell, overflow_direction=None) == ""


def test_meal_sheet_preview_exposes_row_and_overflow_attributes(client, app):
    accountant_headers = _login(client, "accountant")
    category_id = _category_id_by_code(app, "ovz")

    response = client.post(
        "/api/reports/accounting-documents/meal-sheet/document",
        headers=accountant_headers,
        json={
            "month": 2,
            "year": 2025,
            "category_id": category_id,
            "meal_type": "breakfast",
        },
    )

    assert response.status_code == 200
    html = response.get_json()["data"]["html"]

    assert 'data-accounting-row="7"' in html
    assert 'data-accounting-row-height-mm="' in html
    assert re.search(
        r'data-accounting-cell="AR7"[^>]*data-accounting-overflow="left"|'
        r'data-accounting-overflow="left"[^>]*data-accounting-cell="AR7"',
        html,
    )


def test_meal_sheet_preview_as7_left_overflow_uses_absolute_right_anchor(client, app):
    accountant_headers = _login(client, "accountant")
    category_id = _category_id_by_code(app, "large_family")

    response = client.post(
        "/api/reports/accounting-documents/meal-sheet/document",
        headers=accountant_headers,
        json={
            "month": 2,
            "year": 2025,
            "category_id": category_id,
            "meal_type": "breakfast",
        },
    )

    assert response.status_code == 200
    html = response.get_json()["data"]["html"]
    cell_html_match = re.search(
        r'<td(?=[^>]*data-accounting-cell="AS7")(?=[^>]*data-accounting-overflow="left")[^>]*>.*?</td>',
        html,
        re.DOTALL,
    )

    assert cell_html_match is not None
    cell_html = cell_html_match.group(0)
    assert '<span style="' in cell_html
    assert "position:absolute" in cell_html
    assert "right:0" in cell_html
    assert "margin-left:auto" not in cell_html


def test_preview_right_aligned_cell_without_overflow_has_no_content_span_style():
    workbook = Workbook()
    sheet = workbook.active
    cell = sheet["A1"]
    cell.value = "100"
    cell.alignment = Alignment(horizontal="right")

    html = _render_worksheet_page(sheet, "A1:A1")

    assert 'data-accounting-cell="A1"' in html
    assert "text-align:right" in html
    assert "data-accounting-overflow" not in html
    assert "<span style=" not in html


def test_preview_print_exposes_scaled_columns_rows_and_fonts():
    workbook = Workbook()
    sheet = workbook.active
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.paperSize = "9"
    sheet.page_setup.fitToWidth = 1
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_margins.left = 0.25
    sheet.page_margins.right = 0.25

    for column_index in range(1, 12):
        sheet.column_dimensions[chr(64 + column_index)].width = 20
        sheet.cell(row=1, column=column_index).value = f"header {column_index}"

    html = _render_worksheet_page(sheet, "A1:K1")

    scale_match = re.search(r'data-print-scale="([0-9.]+)"', html)
    assert scale_match is not None
    assert float(scale_match.group(1)) < 1
    assert "--accounting-screen-height:" in html
    assert "--accounting-print-height:" in html
    assert "--accounting-print-scale:" in html
    assert "--accounting-screen-row-height:" in html
    assert "--accounting-print-row-height:" in html
    assert "--accounting-print-col-width:" in html
    assert "--accounting-cell-font-size:" in html
    assert "--accounting-print-font-size:" in html


def test_preview_print_scale_ignores_height_without_fit_to_height():
    workbook = Workbook()
    sheet = workbook.active
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.paperSize = "9"
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_margins.left = 0.25
    sheet.page_margins.right = 0.25
    sheet.page_margins.top = 0.25
    sheet.page_margins.bottom = 0.25

    for column_index in range(1, 9):
        sheet.column_dimensions[chr(64 + column_index)].width = 20
    for row_index in range(1, 121):
        sheet.cell(row=row_index, column=1).value = f"row {row_index}"

    width_only_html = _render_worksheet_page(sheet, "A1:H120")
    width_only_scale = float(re.search(r'data-print-scale="([0-9.]+)"', width_only_html).group(1))

    sheet.page_setup.fitToHeight = 1
    single_page_html = _render_worksheet_page(sheet, "A1:H120")
    single_page_scale = float(re.search(r'data-print-scale="([0-9.]+)"', single_page_html).group(1))

    assert width_only_scale < 1
    assert single_page_scale < width_only_scale
