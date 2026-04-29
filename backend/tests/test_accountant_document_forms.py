from __future__ import annotations

import re
from calendar import monthrange
from datetime import date, timedelta
from html import unescape
from io import BytesIO
from xml.etree import ElementTree

import pytest
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import get_column_letter

from app.models import Category, HolidayCalendar, Student, db
from app.services.accounting_documents.cell_values import format_cell_display_value
from app.services.accounting_documents.meal_sheet_header import resolve_meal_sheet_institution_cell
from app.services.accounting_documents.metadata import (
    ALL_CATEGORIES_COST_CALCULATION_LINE,
    CALCULATION_CATEGORY_LINES,
)
from app.services.accounting_documents.template_config import (
    resolve_cost_calculation_template,
    resolve_cost_statement_template,
    resolve_meal_sheet_template,
)
from app.utils.official_holidays import ensure_official_holidays_for_years


PAPER_SIZES_MM: dict[int, tuple[float, float]] = {
    1: (215.9, 279.4),
    5: (215.9, 355.6),
    8: (297.0, 420.0),
    9: (210.0, 297.0),
}
MM_PER_INCH = 25.4


def login(client, username: str, password: str = "password123") -> dict[str, str]:
    response = client.post("/api/auth/login", json={"username": username, "password": password})
    assert response.status_code == 200
    token = response.get_json()["data"]["token"]
    return {"Authorization": f"Bearer {token}"}


def accounting_service_day(app) -> date:
    current = date.today()
    month_end = current.replace(day=monthrange(current.year, current.month)[1])
    with app.app_context():
        ensure_official_holidays_for_years({current.year})
        holiday_dates = {
            entry.holiday_date
            for entry in HolidayCalendar.query.filter(
                HolidayCalendar.is_active.is_(True),
                HolidayCalendar.holiday_date.between(current, month_end),
            ).all()
        }

    while current.weekday() == 6 or current in holiday_dates:
        current += timedelta(days=1)
    return current


def ensure_student(
    app,
    *,
    category_code: str,
    student_card: str,
    full_name: str,
    group_name: str,
    building_id: int = 1,
) -> None:
    with app.app_context():
        existing_student = Student.query.filter_by(student_card=student_card).first()
        if existing_student is not None:
            return

        category = Category.query.filter_by(code=category_code).first()
        assert category is not None

        db.session.add(
            Student(
                full_name=full_name,
                student_card=student_card,
                group_name=group_name,
                building_id=building_id,
                category_id=category.id,
            )
        )
        db.session.commit()


def prepare_accounting_records(
    client,
    app,
    *,
    student_card: str,
    meal_types: tuple[str, ...],
    category_code: str | None = None,
    full_name: str | None = None,
    group_name: str = "РўР•РЎРў-100",
) -> dict:
    if category_code is not None:
        ensure_student(
            app,
            category_code=category_code,
            student_card=student_card,
            full_name=full_name or f"РЎС‚СѓРґРµРЅС‚ {student_card}",
            group_name=group_name,
        )

    social_headers = login(client, "headsocial")
    cashier_headers = login(client, "cashier1")
    current_day = accounting_service_day(app)

    students_response = client.get(f"/api/students/search?q={student_card}", headers=social_headers)
    assert students_response.status_code == 200
    student = next(item for item in students_response.get_json()["data"] if item["student_card"] == student_card)

    ticket_response = client.post(
        "/api/tickets",
        headers=social_headers,
        json={"student_id": student["id"], "month": current_day.month, "year": current_day.year},
    )
    assert ticket_response.status_code in {201, 409}

    for meal_type in meal_types:
        meal_response = client.post(
            "/api/meals/record",
            headers=cashier_headers,
            json={"code": student["student_card"], "meal_type": meal_type, "issue_date": current_day.isoformat()},
        )
        assert meal_response.status_code in {201, 409}

    return {
        "student": student,
        "month": current_day.month,
        "year": current_day.year,
        "service_day": current_day,
        "category_id": student["category_id"],
    }


def find_day_column(sheet, *, header_row: int, day_number: int) -> str:
    for cell in sheet[header_row]:
        if isinstance(cell, MergedCell):
            continue
        if cell.value == day_number:
            return cell.column_letter
    raise AssertionError(f"РќРµ РЅР°Р№РґРµРЅ СЃС‚РѕР»Р±РµС† РґРЅСЏ {day_number} РІ СЃС‚СЂРѕРєРµ {header_row}")


def visible_day_numbers(sheet, *, header_row: int) -> list[int]:
    values: list[int] = []
    for cell in sheet[header_row]:
        if isinstance(cell, MergedCell):
            continue
        if sheet.column_dimensions[cell.column_letter].hidden:
            continue
        if isinstance(cell.value, int):
            values.append(cell.value)
    return values


def visible_day_widths(sheet, *, header_row: int) -> list[float]:
    widths: list[float] = []
    for cell in sheet[header_row]:
        if isinstance(cell, MergedCell):
            continue
        if sheet.column_dimensions[cell.column_letter].hidden:
            continue
        if isinstance(cell.value, int):
            widths.append(float(sheet.column_dimensions[cell.column_letter].width or 0))
    return widths


def assert_worksheet_columns_follow_contract(sheet, contract: tuple[tuple[str, float], ...], *, tolerance: float = 0.001) -> None:
    for column_letter, expected_width in contract:
        actual_width = float(sheet.column_dimensions[column_letter].width or 0)
        assert abs(actual_width - expected_width) <= tolerance


def extract_width_metrics(html: str) -> tuple[float, float, float]:
    match = re.search(
        (
            r'data-screen-width-mm="(?P<screen>[0-9.]+)".*?'
            r'data-print-width-mm="(?P<width>[0-9.]+)".*?'
            r'data-printable-width-mm="(?P<printable>[0-9.]+)"'
        ),
        html,
    )
    assert match is not None
    return float(match.group("screen")), float(match.group("width")), float(match.group("printable"))


def extract_print_metrics(html: str) -> tuple[float, float]:
    _, print_width, printable_width = extract_width_metrics(html)
    return print_width, printable_width


def extract_screen_col_widths_mm(html: str) -> list[float]:
    return [
        float(width)
        for width in re.findall(r"--accounting-screen-col-width:([0-9.]+)mm;--accounting-print-col-width:", html)
    ]


def worksheet_printable_width_mm(sheet) -> float:
    try:
        page_width_mm, page_height_mm = PAPER_SIZES_MM[int(sheet.page_setup.paperSize)]
    except (KeyError, TypeError, ValueError):
        page_width_mm, page_height_mm = PAPER_SIZES_MM[9]

    if sheet.page_setup.orientation == "landscape":
        width_mm = max(page_width_mm, page_height_mm)
    else:
        width_mm = min(page_width_mm, page_height_mm)

    margins = sheet.page_margins
    return width_mm - ((margins.left or 0.7) + (margins.right or 0.7)) * MM_PER_INCH


def worksheet_visible_width_mm(sheet) -> float:
    total_width = 0.0
    for column_index in range(1, sheet.max_column + 1):
        column_letter = get_column_letter(column_index)
        if sheet.column_dimensions[column_letter].hidden:
            continue

        width = float(sheet.column_dimensions[column_letter].width or 8.43)
        pixels = width * 12 if width < 1 else width * 7 + 5
        total_width += pixels * MM_PER_INCH / 96
    return total_width


def worksheet_visible_column_widths_mm(sheet) -> list[float]:
    widths: list[float] = []
    for column_index in range(1, sheet.max_column + 1):
        column_letter = get_column_letter(column_index)
        if sheet.column_dimensions[column_letter].hidden:
            continue

        width = float(sheet.column_dimensions[column_letter].width or 8.43)
        pixels = width * 12 if width < 1 else width * 7 + 5
        widths.append(pixels * MM_PER_INCH / 96)
    return widths


def assert_preview_and_workbook_column_widths_match(html: str, sheet, *, tolerance_mm: float = 0.2) -> None:
    preview_widths = extract_screen_col_widths_mm(html)
    workbook_widths = worksheet_visible_column_widths_mm(sheet)
    assert len(preview_widths) == len(workbook_widths)
    for preview_width, workbook_width in zip(preview_widths, workbook_widths, strict=True):
        assert abs(preview_width - workbook_width) <= tolerance_mm


def workbook_from_response(response) -> tuple:
    assert response.status_code == 200
    workbook = load_workbook(BytesIO(response.data))
    return workbook, workbook.active


def workbook_cell_display_value(sheet, cell_ref: str) -> str:
    cell = sheet[cell_ref]
    return format_cell_display_value(cell.value, cell.number_format)


def extracted_cell_display_value(
    sheet,
    cell_ref: str,
    *,
    prefix: str = "",
    suffix: str = "",
) -> str:
    value = workbook_cell_display_value(sheet, cell_ref)
    if prefix and value.startswith(prefix):
        value = value[len(prefix) :]
    if suffix and value.endswith(suffix):
        value = value[: -len(suffix)]
    return value.strip()


def editable_metadata_by_key(payload: dict) -> dict[str, dict]:
    return {field["key"]: field for field in payload["editable_metadata"]}


def prepared_by_cell_text(sheet, *, binding) -> str:
    assert binding is not None
    return str(sheet[binding.cell].value or "")


def expected_prepared_by_cell_text(*, binding, short_name: str) -> str:
    assert binding is not None
    if binding.mode == "value_cell":
        return short_name
    return f"{binding.prefix}{short_name}"


def visible_sheet_text(sheet) -> str:
    parts: list[str] = []
    for row in sheet.iter_rows():
        if sheet.row_dimensions[row[0].row].hidden:
            continue
        for cell in row:
            if isinstance(cell, MergedCell):
                continue
            if cell.value in {None, ""}:
                continue
            parts.append(str(cell.value))
    return " ".join(parts).lower()


def visible_row_text_values(sheet, *, row_index: int) -> list[str]:
    if sheet.row_dimensions[row_index].hidden:
        return []

    values: list[str] = []
    for cell in sheet[row_index]:
        if isinstance(cell, MergedCell):
            continue
        if isinstance(cell.value, str) and cell.value.strip():
            values.append(cell.value.strip())
    return values


def visible_footer_text_values(sheet, *, start_row: int) -> list[str]:
    values: list[str] = []
    for row_index in range(start_row, sheet.max_row + 1):
        values.extend(visible_row_text_values(sheet, row_index=row_index))
    return values


def html_table_rows(html: str) -> list[list[dict[str, str | None]]]:
    xml = f"<root>{html}</root>".replace("&nbsp;", " ")
    root = ElementTree.fromstring(xml)
    rows: list[list[dict[str, str | None]]] = []

    for row in root.findall(".//tr"):
        parsed_row: list[dict[str, str | None]] = []
        for cell in row.findall("td"):
            text = " ".join("".join(cell.itertext()).split())
            parsed_row.append(
                {
                    "text": text,
                    "colspan": cell.attrib.get("colspan"),
                    "rowspan": cell.attrib.get("rowspan"),
                }
            )
        rows.append(parsed_row)

    return rows


def find_html_cell(rows: list[list[dict[str, str | None]]], text: str) -> dict[str, str | None]:
    for row in rows:
        for cell in row:
            if cell["text"] == text:
                return cell
    raise AssertionError(f"Cell with text {text!r} not found in HTML")


def test_accountant_meal_sheet_breakfast_document_and_xlsx_use_category_data(client, app):
    accountant_headers = login(client, "accountant")
    prepared = prepare_accounting_records(client, app, student_card="100002", meal_types=("breakfast", "lunch"))
    config = resolve_meal_sheet_template("ovz", "breakfast")
    document_response = client.post(
        "/api/reports/accounting-documents/meal-sheet/document",
        headers=accountant_headers,
        json={
            "month": prepared["month"],
            "year": prepared["year"],
            "category_id": prepared["category_id"],
            "meal_type": "breakfast",
        },
    )

    assert document_response.status_code == 200
    payload = document_response.get_json()["data"]
    assert payload["print_mode"] == "embedded"
    assert payload["page_orientation"] == config.page_orientation
    assert prepared["student"]["full_name"] in payload["html"]
    assert "Стоимость завтрака:" in payload["html"]
    assert ">X<" in payload["html"]

    screen_width, rendered_width, printable_width = extract_width_metrics(payload["html"])
    assert rendered_width <= printable_width + 0.05

    workbook_response = client.post(
        "/api/reports/accounting-documents/meal-sheet/xlsx",
        headers=accountant_headers,
        json={
            "month": prepared["month"],
            "year": prepared["year"],
            "category_id": prepared["category_id"],
            "meal_type": "breakfast",
        },
    )

    _, sheet = workbook_from_response(workbook_response)
    day_column = find_day_column(sheet, header_row=config.header_day_row, day_number=prepared["service_day"].day)

    assert str(sheet[config.title_cell].value).startswith("Накопительная ведомость")
    assert sheet[f"{config.student_column}{config.data_start_row}"].value == prepared["student"]["full_name"]
    assert sheet[f"{config.group_column}{config.data_start_row}"].value == prepared["student"]["group_name"]
    assert sheet[config.price_label_cell].value == "Стоимость завтрака:"
    assert sheet[f"{day_column}{config.data_start_row}"].value == "X"
    assert sheet.row_dimensions[config.data_start_row + 1].hidden is True
    footer_border = sheet[f"E{config.amount_row + 1}"].border
    assert footer_border.left is None or footer_border.left.style is None
    assert footer_border.right is None or footer_border.right.style is None
    signature_row_values = visible_row_text_values(sheet, row_index=config.signature_row)
    assert signature_row_values
    assert any(value in payload["html"] for value in signature_row_values)
    assert_preview_and_workbook_column_widths_match(payload["html"], sheet)
    preview_widths = extract_screen_col_widths_mm(payload["html"])
    assert preview_widths[0] < preview_widths[1]
    assert screen_width > rendered_width + 0.05
    assert abs(screen_width - worksheet_visible_width_mm(sheet)) <= 0.1
    assert rendered_width <= worksheet_printable_width_mm(sheet) + 0.05


def test_accountant_meal_sheet_lunch_document_supports_low_income_category(client, app):
    accountant_headers = login(client, "accountant")
    prepared = prepare_accounting_records(
        client,
        app,
        student_card="200001",
        category_code="low_income",
        full_name="РЎРјРёСЂРЅРѕРІР° РђРЅРЅР° РћР»РµРіРѕРІРЅР°",
        group_name="Р‘РЈРҐ-11",
        meal_types=("lunch",),
    )
    config = resolve_meal_sheet_template("low_income", "lunch")

    document_response = client.post(
        "/api/reports/accounting-documents/meal-sheet/document",
        headers=accountant_headers,
        json={
            "month": prepared["month"],
            "year": prepared["year"],
            "category_id": prepared["category_id"],
            "meal_type": "lunch",
        },
    )

    assert document_response.status_code == 200
    payload = document_response.get_json()["data"]
    assert prepared["student"]["full_name"] in payload["html"]
    assert "Стоимость обеда:" in payload["html"]

    workbook_response = client.post(
        "/api/reports/accounting-documents/meal-sheet/xlsx",
        headers=accountant_headers,
        json={
            "month": prepared["month"],
            "year": prepared["year"],
            "category_id": prepared["category_id"],
            "meal_type": "lunch",
        },
    )

    _, sheet = workbook_from_response(workbook_response)
    day_column = find_day_column(sheet, header_row=config.header_day_row, day_number=prepared["service_day"].day)

    assert "малоимущ" in str(sheet[config.title_cell].value).lower()
    assert sheet[f"{config.student_column}{config.data_start_row}"].value == prepared["student"]["full_name"]
    assert sheet[f"{day_column}{config.data_start_row}"].value == "X"
    assert sheet[f"{config.total_column}{config.data_start_row}"].value in {150, 150.0}


def test_accountant_meal_sheet_lunch_preview_keeps_number_and_day_columns_stable(client, app):
    accountant_headers = login(client, "accountant")
    prepared = prepare_accounting_records(client, app, student_card="100002", meal_types=("breakfast", "lunch"))
    config = resolve_meal_sheet_template("ovz", "lunch")

    document_response = client.post(
        "/api/reports/accounting-documents/meal-sheet/document",
        headers=accountant_headers,
        json={
            "month": prepared["month"],
            "year": prepared["year"],
            "category_id": prepared["category_id"],
            "meal_type": "lunch",
        },
    )
    assert document_response.status_code == 200
    payload = document_response.get_json()["data"]

    workbook_response = client.post(
        "/api/reports/accounting-documents/meal-sheet/xlsx",
        headers=accountant_headers,
        json={
            "month": prepared["month"],
            "year": prepared["year"],
            "category_id": prepared["category_id"],
            "meal_type": "lunch",
        },
    )
    _, sheet = workbook_from_response(workbook_response)

    assert_preview_and_workbook_column_widths_match(payload["html"], sheet)
    preview_widths = extract_screen_col_widths_mm(payload["html"])

    assert len(preview_widths) >= 13
    assert preview_widths[0] < preview_widths[11]
    assert abs(preview_widths[10] - preview_widths[11]) < 0.01
    assert abs(preview_widths[11] - preview_widths[12]) < 0.01
    assert sheet[config.title_cell].value is not None


def test_accountant_meal_sheet_disabled_category_overrides_header_and_hides_technical_columns(client, app):
    accountant_headers = login(client, "accountant")
    prepared = prepare_accounting_records(
        client,
        app,
        student_card="200003",
        category_code="disabled",
        full_name="РџРµС‚СЂРѕРІ РРІР°РЅ РЎРµСЂРіРµРµРІРёС‡",
        group_name="РСЂ3-18",
        meal_types=("lunch",),
    )
    config = resolve_meal_sheet_template("disabled", "lunch")

    workbook_response = client.post(
        "/api/reports/accounting-documents/meal-sheet/xlsx",
        headers=accountant_headers,
        json={
            "month": prepared["month"],
            "year": prepared["year"],
            "category_id": prepared["category_id"],
            "meal_type": "lunch",
        },
    )

    _, sheet = workbook_from_response(workbook_response)

    assert "инвалид" in str(sheet[config.title_cell].value).lower()
    assert "инвалид" in str(sheet[config.student_header_cell].value).lower()
    assert all(sheet.column_dimensions[column].hidden is True for column in ("Y", "Z", "AA", "AB", "AC", "AD"))


def test_accountant_cost_statement_document_and_xlsx_hide_top_signature_block_and_keep_codes_aligned(client, app):
    accountant_headers = login(client, "accountant")
    prepared = prepare_accounting_records(
        client,
        app,
        student_card="200001",
        category_code="low_income",
        full_name="РЎРјРёСЂРЅРѕРІР° РђРЅРЅР° РћР»РµРіРѕРІРЅР°",
        group_name="Р‘РЈРҐ-11",
        meal_types=("lunch",),
    )
    config = resolve_cost_statement_template("low_income")

    document_response = client.post(
        "/api/reports/accounting-documents/cost-statement/document",
        headers=accountant_headers,
        json={
            "month": prepared["month"],
            "year": prepared["year"],
            "category_id": prepared["category_id"],
        },
    )

    assert document_response.status_code == 200
    payload = document_response.get_json()["data"]
    assert payload["print_mode"] == "embedded"
    assert payload["page_orientation"] == config.page_orientation
    assert prepared["student"]["full_name"] in payload["html"]
    assert "Р”РёСЂРµРєС‚РѕСЂ" not in payload["html"]
    assert "Р“Р». Р±СѓС…РіР°Р»С‚РµСЂ" not in payload["html"]

    html_rows = html_table_rows(payload["html"])
    assert find_html_cell(html_rows, "Форма 389 по ОКУД")["colspan"] is None
    assert find_html_cell(html_rows, "по ОКПО")["colspan"] is None
    assert find_html_cell(html_rows, "по ОКЕИ")["colspan"] is None
    assert "'Times New Roman', serif" in unescape(payload["html"])

    rendered_width, printable_width = extract_print_metrics(payload["html"])
    assert rendered_width <= printable_width + 0.05

    workbook_response = client.post(
        "/api/reports/accounting-documents/cost-statement/xlsx",
        headers=accountant_headers,
        json={
            "month": prepared["month"],
            "year": prepared["year"],
            "category_id": prepared["category_id"],
        },
    )

    _, sheet = workbook_from_response(workbook_response)
    assert_worksheet_columns_follow_contract(sheet, config.column_width_contract)

    for row_index in range(1, 6):
        assert sheet.row_dimensions[row_index].hidden is True

    assert sheet[f"{config.student_column}{config.data_start_row}"].value == prepared["student"]["full_name"]
    assert sheet[f"{config.amount_column}{config.data_start_row}"].value in {150, 150.0}
    assert sheet.max_row >= config.total_row
    footer_values = visible_footer_text_values(sheet, start_row=config.total_row + 1)
    assert footer_values
    assert any(value in payload["html"] for value in footer_values)
    assert rendered_width <= worksheet_printable_width_mm(sheet) + 0.05
    assert_preview_and_workbook_column_widths_match(payload["html"], sheet)


def test_accountant_cost_statement_preview_column_widths_match_workbook(client, app):
    accountant_headers = login(client, "accountant")
    config = resolve_cost_statement_template("low_income")
    prepared = prepare_accounting_records(
        client,
        app,
        student_card="200001",
        category_code="low_income",
        full_name="Р РЋР СР С‘РЎР‚Р Р…Р С•Р Р†Р В° Р С’Р Р…Р Р…Р В° Р С›Р В»Р ВµР С–Р С•Р Р†Р Р…Р В°",
        group_name="Р вЂР Р€Р Тђ-11",
        meal_types=("lunch",),
    )

    document_response = client.post(
        "/api/reports/accounting-documents/cost-statement/document",
        headers=accountant_headers,
        json={
            "month": prepared["month"],
            "year": prepared["year"],
            "category_id": prepared["category_id"],
        },
    )
    assert document_response.status_code == 200
    payload = document_response.get_json()["data"]

    workbook_response = client.post(
        "/api/reports/accounting-documents/cost-statement/xlsx",
        headers=accountant_headers,
        json={
            "month": prepared["month"],
            "year": prepared["year"],
            "category_id": prepared["category_id"],
        },
    )
    _, sheet = workbook_from_response(workbook_response)

    assert_preview_and_workbook_column_widths_match(payload["html"], sheet)
    assert_worksheet_columns_follow_contract(sheet, config.column_width_contract)
    preview_widths = extract_screen_col_widths_mm(payload["html"])
    assert preview_widths[0] < preview_widths[1]


def test_accountant_cost_statement_uses_disabled_category_caption(client, app):
    accountant_headers = login(client, "accountant")
    prepared = prepare_accounting_records(
        client,
        app,
        student_card="200004",
        category_code="disabled",
        full_name="РЎРѕРєРѕР»РѕРІР° РњР°СЂРёРЅР° РРіРѕСЂРµРІРЅР°",
        group_name="Р®Р -21",
        meal_types=("lunch",),
    )
    config = resolve_cost_statement_template("disabled")

    workbook_response = client.post(
        "/api/reports/accounting-documents/cost-statement/xlsx",
        headers=accountant_headers,
        json={
            "month": prepared["month"],
            "year": prepared["year"],
            "category_id": prepared["category_id"],
        },
    )

    _, sheet = workbook_from_response(workbook_response)

    assert sheet[config.category_cell].value == 'категория "инвалиды"'


@pytest.mark.parametrize(
    ("student_card", "meal_types", "category_code", "expected_line", "expected_meal_amount"),
    [
        ("100001", ("lunch",), "orphan", CALCULATION_CATEGORY_LINES["orphan"], 175.0),
        ("100002", ("breakfast", "lunch"), "ovz", CALCULATION_CATEGORY_LINES["ovz"], 260.0),
        ("200005", ("lunch",), "disabled", 'студентам категории "инвалиды" колледжа', 165.0),
    ],
)
def test_accountant_cost_calculation_document_and_xlsx_use_dynamic_category_line(
    client,
    app,
    student_card: str,
    meal_types: tuple[str, ...],
    category_code: str,
    expected_line: str,
    expected_meal_amount: float,
):
    accountant_headers = login(client, "accountant")
    prepared = prepare_accounting_records(
        client,
        app,
        student_card=student_card,
        meal_types=meal_types,
        category_code=category_code,
    )
    config = resolve_cost_calculation_template(category_code)

    document_response = client.post(
        "/api/reports/accounting-documents/cost-calculation/document",
        headers=accountant_headers,
        json={
            "month": prepared["month"],
            "year": prepared["year"],
            "category_id": prepared["category_id"],
        },
    )

    assert document_response.status_code == 200
    payload = document_response.get_json()["data"]
    assert payload["page_orientation"] == config.page_orientation
    assert prepared["student"]["full_name"] in payload["html"]
    assert expected_line in unescape(payload["html"])

    workbook_response = client.post(
        "/api/reports/accounting-documents/cost-calculation/xlsx",
        headers=accountant_headers,
        json={
            "month": prepared["month"],
            "year": prepared["year"],
            "category_id": prepared["category_id"],
        },
    )

    _, sheet = workbook_from_response(workbook_response)
    assert_worksheet_columns_follow_contract(sheet, config.column_width_contract)

    assert sheet[config.category_line_cell].value == expected_line
    assert sheet[f"{config.student_column}{config.data_start_row}"].value == prepared["student"]["full_name"]
    assert sheet[f"{config.meal_amount_column}{config.data_start_row}"].value in {expected_meal_amount, int(expected_meal_amount)}
    assert sheet[f"{config.meal_amount_column}{config.total_row}"].value in {expected_meal_amount, int(expected_meal_amount)}
    signature_row_values = visible_row_text_values(sheet, row_index=config.signature_row)
    assert signature_row_values
    assert any(value in payload["html"] for value in signature_row_values)


def test_accountant_all_categories_documents_aggregate_supported_categories(client, app):
    accountant_headers = login(client, "accountant")
    orphan_prepared = prepare_accounting_records(client, app, student_card="100001", meal_types=("lunch",))
    ovz_prepared = prepare_accounting_records(client, app, student_card="100002", meal_types=("breakfast", "lunch"))
    low_income_prepared = prepare_accounting_records(
        client,
        app,
        student_card="200001",
        category_code="low_income",
        full_name="РЎРјРёСЂРЅРѕРІР° РђРЅРЅР° РћР»РµРіРѕРІРЅР°",
        group_name="Р‘Р¦-11",
        meal_types=("lunch",),
    )
    large_family_prepared = prepare_accounting_records(
        client,
        app,
        student_card="200002",
        category_code="large_family",
        full_name="РљСѓР·РЅРµС†РѕРІР° РњР°СЂРёСЏ РРІР°РЅРѕРІРЅР°",
        group_name="РРЎРџ-44",
        meal_types=("breakfast", "lunch"),
    )
    lunch_names = {
        orphan_prepared["student"]["full_name"],
        ovz_prepared["student"]["full_name"],
        low_income_prepared["student"]["full_name"],
        large_family_prepared["student"]["full_name"],
    }

    meal_document_response = client.post(
        "/api/reports/accounting-documents/meal-sheet/document",
        headers=accountant_headers,
        json={
            "month": orphan_prepared["month"],
            "year": orphan_prepared["year"],
            "category_id": 0,
            "meal_type": "lunch",
        },
    )
    assert meal_document_response.status_code == 200
    meal_payload = meal_document_response.get_json()["data"]
    for student_name in lunch_names:
        assert student_name in meal_payload["html"]

    meal_workbook_response = client.post(
        "/api/reports/accounting-documents/meal-sheet/xlsx",
        headers=accountant_headers,
        json={
            "month": orphan_prepared["month"],
            "year": orphan_prepared["year"],
            "category_id": 0,
            "meal_type": "lunch",
        },
    )
    _, meal_sheet = workbook_from_response(meal_workbook_response)
    assert "всех категорий" in str(meal_sheet["A7"].value).lower()

    workbook_lunch_names = {
        meal_sheet["B12"].value,
        meal_sheet["B13"].value,
        meal_sheet["B14"].value,
        meal_sheet["B15"].value,
    }
    assert workbook_lunch_names == lunch_names
    meal_signature_row_values = visible_row_text_values(
        meal_sheet,
        row_index=resolve_meal_sheet_template("all", "lunch").signature_row,
    )
    assert meal_signature_row_values
    assert any(value in meal_payload["html"] for value in meal_signature_row_values)

    statement_document_response = client.post(
        "/api/reports/accounting-documents/cost-statement/document",
        headers=accountant_headers,
        json={
            "month": orphan_prepared["month"],
            "year": orphan_prepared["year"],
            "category_id": 0,
        },
    )
    assert statement_document_response.status_code == 200
    statement_payload = statement_document_response.get_json()["data"]
    for student_name in lunch_names:
        assert student_name in statement_payload["html"]

    statement_config = resolve_cost_statement_template("all")
    statement_workbook_response = client.post(
        "/api/reports/accounting-documents/cost-statement/xlsx",
        headers=accountant_headers,
        json={
            "month": orphan_prepared["month"],
            "year": orphan_prepared["year"],
            "category_id": 0,
        },
    )
    _, statement_sheet = workbook_from_response(statement_workbook_response)
    assert_worksheet_columns_follow_contract(statement_sheet, statement_config.column_width_contract)
    assert statement_sheet[f"{statement_config.amount_column}{statement_config.total_row}"].value in {830, 830.0}
    statement_footer_values = visible_footer_text_values(statement_sheet, start_row=statement_config.total_row + 1)
    assert statement_footer_values
    assert any(value in statement_payload["html"] for value in statement_footer_values)

    calculation_config = resolve_cost_calculation_template("all")
    calculation_workbook_response = client.post(
        "/api/reports/accounting-documents/cost-calculation/xlsx",
        headers=accountant_headers,
        json={
            "month": orphan_prepared["month"],
            "year": orphan_prepared["year"],
            "category_id": 0,
        },
    )
    _, calculation_sheet = workbook_from_response(calculation_workbook_response)
    assert_worksheet_columns_follow_contract(calculation_sheet, calculation_config.column_width_contract)
    assert calculation_sheet[calculation_config.category_line_cell].value == ALL_CATEGORIES_COST_CALCULATION_LINE
    assert calculation_sheet[f"{calculation_config.meal_amount_column}{calculation_config.total_row}"].value in {830, 830.0}


def test_accountant_meal_sheet_hides_unused_days_and_rows_for_february(client):
    accountant_headers = login(client, "accountant")
    config = resolve_meal_sheet_template("ovz", "breakfast")

    response = client.post(
        "/api/reports/accounting-documents/meal-sheet/xlsx",
        headers=accountant_headers,
        json={
            "month": 2,
            "year": 2025,
            "category_id": 1,
            "meal_type": "breakfast",
        },
    )

    _, sheet = workbook_from_response(response)

    assert visible_day_numbers(sheet, header_row=config.header_day_row) == list(range(1, 29))
    assert sheet.row_dimensions[config.data_start_row].hidden is True
    assert sheet.row_dimensions[config.data_start_row + 1].hidden is True


@pytest.mark.parametrize(
    ("category_id", "category_code", "meal_type"),
    [
        (1, "ovz", "breakfast"),
        (5, "low_income", "lunch"),
        (2, "orphan", "lunch"),
    ],
)
def test_accountant_meal_sheet_normalizes_day_column_widths(client, category_id: int, category_code: str, meal_type: str):
    accountant_headers = login(client, "accountant")
    config = resolve_meal_sheet_template(category_code, meal_type)

    response = client.post(
        "/api/reports/accounting-documents/meal-sheet/xlsx",
        headers=accountant_headers,
        json={
            "month": 2,
            "year": 2025,
            "category_id": category_id,
            "meal_type": meal_type,
        },
    )

    _, sheet = workbook_from_response(response)
    widths = visible_day_widths(sheet, header_row=config.header_day_row)

    assert len(widths) == 28
    assert max(widths) - min(widths) < 0.01
    assert config.day_slot_width is not None
    assert all(abs(width - config.day_slot_width) < 0.01 for width in widths)


@pytest.mark.parametrize(
    ("category_code", "meal_type", "template_cell"),
    [
        ("low_income", "lunch", "AP7"),
        ("ovz", "breakfast", "AR7"),
        ("ovz", "lunch", "AO7"),
        ("orphan", "lunch", "AQ1"),
        ("large_family", "breakfast", "AS7"),
        ("large_family", "lunch", "AP7"),
    ],
)
def test_accountant_meal_sheet_institution_binding_uses_source_cell(
    client,
    app,
    category_code: str,
    meal_type: str,
    template_cell: str,
):
    accountant_headers = login(client, "accountant")
    config = resolve_meal_sheet_template(category_code, meal_type)
    with app.app_context():
        category = Category.query.filter_by(code=category_code).first()
        assert category is not None
        category_id = int(category.id)

    document_response = client.post(
        "/api/reports/accounting-documents/meal-sheet/document",
        headers=accountant_headers,
        json={
            "month": 2,
            "year": 2025,
            "category_id": category_id,
            "meal_type": meal_type,
        },
    )
    workbook_response = client.post(
        "/api/reports/accounting-documents/meal-sheet/xlsx",
        headers=accountant_headers,
        json={
            "month": 2,
            "year": 2025,
            "category_id": category_id,
            "meal_type": meal_type,
        },
    )

    assert document_response.status_code == 200
    assert workbook_response.status_code == 200

    payload = document_response.get_json()["data"]
    metadata = editable_metadata_by_key(payload)
    _, sheet = workbook_from_response(workbook_response)
    institution_cell = resolve_meal_sheet_institution_cell(sheet, config)

    assert config.institution_cell == template_cell
    assert institution_cell == template_cell
    assert metadata["institution"]["cell"] == template_cell
    assert metadata["institution"]["value"] == workbook_cell_display_value(sheet, template_cell)
    assert workbook_cell_display_value(sheet, template_cell) == "МЦК - ЧЭМК Минобразования Чувашии"
    assert f'data-accounting-cell="{template_cell}"' in payload["html"]



def test_accountant_cost_statement_document_exposes_cell_bound_editable_metadata(client, app):
    accountant_headers = login(client, "accountant")
    prepared = prepare_accounting_records(
        client,
        app,
        student_card="200011",
        category_code="low_income",
        full_name="РџРѕР»СЏРєРѕРІР° РњР°СЂРёРЅР° РЎРµСЂРіРµРµРІРЅР°",
        group_name="Р‘РЈРҐ-21",
        meal_types=("lunch",),
    )
    config = resolve_cost_statement_template("low_income")

    document_response = client.post(
        "/api/reports/accounting-documents/cost-statement/document",
        headers=accountant_headers,
        json={
            "month": prepared["month"],
            "year": prepared["year"],
            "category_id": prepared["category_id"],
        },
    )

    assert document_response.status_code == 200
    payload = document_response.get_json()["data"]
    metadata = editable_metadata_by_key(payload)
    workbook_response = client.post(
        "/api/reports/accounting-documents/cost-statement/xlsx",
        headers=accountant_headers,
        json={
            "month": prepared["month"],
            "year": prepared["year"],
            "category_id": prepared["category_id"],
        },
    )

    _, sheet = workbook_from_response(workbook_response)
    checked_row = int("".join(char for char in config.prepared_by_binding.cell if char.isdigit())) + 2

    assert f'data-accounting-cell="{config.institution_cell}"' in payload["html"]
    assert set(metadata) == {
        "title",
        "subtitle",
        "okudForm",
        "okudCode",
        "monthLabel",
        "yearLabel",
        "reportDate",
        "institution",
        "structuralUnit",
        "division",
        "targetArticle",
        "expenseType",
        "measurementUnit",
        "fundingSource",
        "okpoCode",
        "kspCode",
        "fkrCode",
        "kcsrCode",
        "kvrCode",
        "okeiCode",
        "preparedByLabel",
        "preparedByName",
        "checkedByLabel",
        "checkedByName",
    }
    assert metadata["title"]["value"] == workbook_cell_display_value(sheet, "E6")
    assert metadata["subtitle"]["value"] == workbook_cell_display_value(sheet, "B7")
    assert metadata["okudForm"]["value"] == extracted_cell_display_value(
        sheet,
        "H7",
        prefix="Форма ",
        suffix=" по ОКУД",
    )
    assert metadata["okudCode"]["value"] == workbook_cell_display_value(sheet, "I7")
    assert metadata["monthLabel"]["value"] == workbook_cell_display_value(sheet, config.month_cell)
    assert metadata["yearLabel"]["value"] == workbook_cell_display_value(sheet, config.year_cell)
    assert metadata["reportDate"]["value"] == workbook_cell_display_value(sheet, config.report_date_cell)
    assert metadata["institution"]["value"] == workbook_cell_display_value(sheet, config.institution_cell)
    assert metadata["structuralUnit"]["value"] == workbook_cell_display_value(sheet, config.category_cell)
    assert metadata["division"]["value"] == workbook_cell_display_value(sheet, "C11")
    assert metadata["targetArticle"]["value"] == workbook_cell_display_value(sheet, "C12")
    assert metadata["expenseType"]["value"] == workbook_cell_display_value(sheet, "C13")
    assert metadata["measurementUnit"]["value"] == "руб."
    assert metadata["fundingSource"]["value"] == workbook_cell_display_value(sheet, "C15")
    assert metadata["okpoCode"]["value"] == workbook_cell_display_value(sheet, "I9")
    assert metadata["kspCode"]["value"] == workbook_cell_display_value(sheet, "I10")
    assert metadata["fkrCode"]["value"] == workbook_cell_display_value(sheet, "I11")
    assert metadata["kcsrCode"]["value"] == workbook_cell_display_value(sheet, "I12")
    assert metadata["kvrCode"]["value"] == workbook_cell_display_value(sheet, "I13")
    assert metadata["okeiCode"]["value"] == workbook_cell_display_value(sheet, "I14")
    assert metadata["preparedByLabel"]["value"] == workbook_cell_display_value(sheet, f"A{checked_row - 2}")
    assert metadata["preparedByName"]["value"] == workbook_cell_display_value(sheet, config.prepared_by_binding.cell)
    assert metadata["checkedByLabel"]["value"] == workbook_cell_display_value(sheet, f"A{checked_row}")
    assert metadata["checkedByName"]["value"] == workbook_cell_display_value(sheet, f"E{checked_row}")


def test_accountant_meal_sheet_document_exposes_editable_metadata(client, app):
    accountant_headers = login(client, "accountant")
    config = resolve_meal_sheet_template("ovz", "breakfast")
    prepared = prepare_accounting_records(
        client,
        app,
        student_card="100012",
        category_code="ovz",
        full_name="РњРёСЂРѕРЅРѕРІР° РђР»С‘РЅР° РРіРѕСЂРµРІРЅР°",
        group_name="РћР’Р—-12",
        meal_types=("breakfast", "lunch"),
    )

    document_response = client.post(
        "/api/reports/accounting-documents/meal-sheet/document",
        headers=accountant_headers,
        json={
            "month": prepared["month"],
            "year": prepared["year"],
            "category_id": prepared["category_id"],
            "meal_type": "breakfast",
        },
    )

    assert document_response.status_code == 200
    payload = document_response.get_json()["data"]
    workbook_response = client.post(
        "/api/reports/accounting-documents/meal-sheet/xlsx",
        headers=accountant_headers,
        json={
            "month": prepared["month"],
            "year": prepared["year"],
            "category_id": prepared["category_id"],
            "meal_type": "breakfast",
        },
    )
    _, sheet = workbook_from_response(workbook_response)
    metadata = editable_metadata_by_key(payload)
    institution_cell = resolve_meal_sheet_institution_cell(sheet, config)

    assert set(metadata) == {
        "title",
        "monthLabel",
        "yearLabel",
        "institution",
        "studentHeader",
        "priceLabel",
        "preparedByName",
    }
    assert metadata["title"]["value"] == workbook_cell_display_value(sheet, config.title_cell)
    assert metadata["monthLabel"]["value"] == workbook_cell_display_value(sheet, config.month_cell)
    assert metadata["yearLabel"]["value"] == workbook_cell_display_value(sheet, config.year_cell)
    assert metadata["institution"]["cell"] == institution_cell
    assert metadata["institution"]["value"] == workbook_cell_display_value(sheet, institution_cell)
    assert metadata["studentHeader"]["value"] == workbook_cell_display_value(sheet, config.student_header_cell)
    assert metadata["priceLabel"]["value"] == workbook_cell_display_value(sheet, config.price_label_cell)
    assert metadata["preparedByName"]["value"] == extracted_cell_display_value(
        sheet,
        config.prepared_by_binding.cell,
        prefix=config.prepared_by_binding.prefix,
    )
    assert f'data-accounting-cell="{institution_cell}"' in payload["html"]
    assert "МЦК - ЧЭМК Минобразования Чувашии" in payload["html"]


def test_accountant_cost_calculation_document_exposes_editable_metadata(client, app):
    accountant_headers = login(client, "accountant")
    config = resolve_cost_calculation_template("orphan")
    prepared = prepare_accounting_records(
        client,
        app,
        student_card="100013",
        category_code="orphan",
        full_name="РЎРѕРєРѕР»РѕРІ РђСЂС‚С‘Рј Р’Р°Р»РµСЂСЊРµРІРёС‡",
        group_name="РЎРР -31",
        meal_types=("lunch",),
    )

    document_response = client.post(
        "/api/reports/accounting-documents/cost-calculation/document",
        headers=accountant_headers,
        json={
            "month": prepared["month"],
            "year": prepared["year"],
            "category_id": prepared["category_id"],
        },
    )

    assert document_response.status_code == 200
    payload = document_response.get_json()["data"]
    workbook_response = client.post(
        "/api/reports/accounting-documents/cost-calculation/xlsx",
        headers=accountant_headers,
        json={
            "month": prepared["month"],
            "year": prepared["year"],
            "category_id": prepared["category_id"],
        },
    )
    _, sheet = workbook_from_response(workbook_response)
    metadata = editable_metadata_by_key(payload)

    assert set(metadata) == {
        "institution",
        "title",
        "subtitle",
        "categoryLine",
        "monthLabel",
        "yearLabel",
        "preparedByName",
    }
    assert 'data-accounting-cell="A1"' in payload["html"]
    assert metadata["institution"]["value"] == workbook_cell_display_value(sheet, "A1")
    assert metadata["title"]["value"] == workbook_cell_display_value(sheet, "A2")
    assert metadata["subtitle"]["value"] == workbook_cell_display_value(sheet, "A3")
    assert metadata["categoryLine"]["value"] == workbook_cell_display_value(sheet, config.category_line_cell)
    assert metadata["monthLabel"]["value"] == workbook_cell_display_value(sheet, config.month_cell)
    assert metadata["yearLabel"]["value"] == workbook_cell_display_value(sheet, config.year_cell)
    assert metadata["preparedByName"]["value"] == extracted_cell_display_value(
        sheet,
        config.prepared_by_binding.cell,
        prefix=config.prepared_by_binding.prefix,
    )


def test_accountant_document_requires_category(client):
    accountant_headers = login(client, "accountant")

    response = client.post(
        "/api/reports/accounting-documents/cost-statement/document",
        headers=accountant_headers,
        json={"month": date.today().month, "year": date.today().year},
    )

    assert response.status_code == 400
    assert "category_id" in response.get_json()["error"]

